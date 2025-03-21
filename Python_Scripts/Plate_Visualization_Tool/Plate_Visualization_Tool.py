# -*- coding: utf-8 -*-
"""
@author: TaliaAlbert
"""
import os
import sys
import datetime
import logging
import shutil
import pandas as pd
import itertools
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl import Workbook


import csv

import math
import time

parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
sys.path.insert(0, parent_dir)
import utils

"""
# Get directories
current_dir, archive_dir, data_files_dir, input_dir, output_dir = utils.setup_dir(create_output_dir = 'y')

# Format the date as 'YYYY-MM-DD'
formatted_date = datetime.date.today().strftime('%Y%m%d')

# Setup logging functionality
log_file_name = f'{formatted_date}_Plate_Visualization_Tool_log.log'
utils.setup_logging(archive_dir, log_filename=log_file_name)


# Select Verso Plate Map
verso_num_files = int(input("Enter the number of Verso Plate Map files: "))
verso_plate_map_df = pd.DataFrame()  # Initialize an empty DataFrame

for i in range(verso_num_files):
    verso_message = f"Select the {i + 1} Plate Map file from the Verso (should end in \".txt\")"
    verso_plate_map_dir = utils.select_file_from_input_dir(input_dir, message=verso_message)
    verso_plate_map_file_name = os.path.basename(verso_plate_map_dir)

    try:
        # Read the file and add a 'Source' column
        df = pd.read_csv(verso_plate_map_dir, sep='\t')
        df['Verso_Source_File'] = verso_plate_map_file_name
        
        # Append data directly to the main DataFrame
        verso_plate_map_df = pd.concat([verso_plate_map_df, df], ignore_index=True)
    except Exception as e:
        print(f"Error reading file {i + 1}: {e}")


# Convert Barcode as Float64 to strings without a ".0", and convert column to string
verso_plate_map_df['Barcode'] = verso_plate_map_df['Barcode'].fillna(-1).astype('Int64').astype(str)  # Convert int to string
verso_plate_map_df = verso_plate_map_df[verso_plate_map_df['Barcode'] != '-1']

verso_plate_map_df['Column'] = verso_plate_map_df['Column'].astype(str)
verso_plate_map_df['Well_id'] = verso_plate_map_df['Row'].astype(str) + verso_plate_map_df['Column']




# Select Worklist file
worklist_num_files = int(input("Enter the number of Worklist files: "))
worklist_df = pd.DataFrame()  # Initialize an empty DataFrame

for i in range(worklist_num_files):
    worklist_message = "Select the worklist associated with this plate map file (should end in \".csv\")"
    worklist_file = utils.select_file_from_input_dir(input_dir, message=worklist_message)
    worklist_file_name = os.path.basename(worklist_file)  # Extract the filename

    try:
        # Read the file as a DataFrame and add a 'Source' column
        df = pd.read_csv(worklist_file, sep=',')
        df['Worklist_Source_File'] = worklist_file_name  # Add file source column

        # Append to the main DataFrame
        worklist_df = pd.concat([worklist_df, df], ignore_index=True)
    except Exception as e:
        print(f"Error reading file {i + 1}: {e}")
"""


# File paths
map_template_path = ""

worklist_file_path = ""
verso_file_path = ""

verso_file_name = os.path.basename(verso_file_path)
worklist_file_name = os.path.basename(worklist_file_path)

# Load the verso file (20250221 PRIM 1.txt)
verso_df = pd.read_csv(verso_file_path, sep="\t")
# Sort by row and then column, to ensure it is A1, A2, ..., H11, H12
verso_df = verso_df.sort_values(by=['Row', 'Column'], ascending=[True, True])

"""
# I think we actually want to keep empty barcode entries, so ignoring for now
# verso_df = verso_df.dropna(subset=['Barcode'])
"""
verso_df['Barcode'] = verso_df['Barcode'].fillna(-1).astype('Int64').astype(str)  # Convert int to string
"""
# I think we want to ignore this too
verso_df = verso_df[verso_df['Barcode'] != '-1']
"""
row_mapping = {'A': 1, 'B': 3, 'C': 5, 'D': 7, 'E': 9, 'F': 11, 'G': 13, 'H': 15}
excel_row_offset = 3
excel_column_offset = 1
verso_df['Excel_Row'] = verso_df['Row'].map(row_mapping) + excel_row_offset
verso_df['Excel_Column'] = verso_df['Column'] + excel_column_offset




# Load the worklist mapping file (2025-02-24 5-HT6 5-HT7A M4 PRIM.csv)
worklist_df = pd.read_csv(worklist_file_path)
worklist_df["CMPD"] = worklist_df["CMPD"].astype(str)

# Ensure required columns exist
required_columns = {"CMPD", "PDSP"}
if not required_columns.issubset(worklist_df.columns):
    raise ValueError(f"Missing required columns in worklist file: {required_columns - set(worklist_df.columns)}")


# Merge barcode_df with worklist_df to associate Barcodes with PDSP numbers
verso_worklist_df = verso_df.merge(worklist_df, left_on="Barcode", right_on="CMPD", how="left")

# Drop 'CMPD' as it's redundant
verso_worklist_df.drop(columns=['CMPD'], inplace=True)



# Identify receptor columns (excluding 'PDSP' and 'Barcode')
receptor_columns = [col for col in worklist_df.columns if col not in {"CMPD", "PDSP", "Assay", "Ref"}]


# Group by Well_ID to aggregate receptors properly
verso_worklist_df["Receptor_Info"] = verso_worklist_df[receptor_columns].apply(
    lambda row: [f"{row[rec]}" for rec in receptor_columns if pd.notna(row[rec])], axis=1
)

# Check if 'Ref' exists before including it in the aggregation
groupby_columns = {
    "Barcode": "first",  # Keep first barcode (same per well)
    "PDSP": "first",  # Keep first PDSP value
    "Receptor_Info": lambda x: list(set(sum(x, [])))  # Merge unique receptor values
}

# Add 'Ref' if it's present in the DataFrame
if "Ref" in verso_worklist_df.columns:
    groupby_columns["Ref"] = "first"

# Perform the groupby operation
well_grouped = verso_worklist_df.groupby("Barcode", as_index=False).agg(groupby_columns)

"""
# Replace NaN 'Ref' and PDSP values with an empty string
well_grouped["Ref"] = well_grouped["Ref"].fillna("")
well_grouped["PDSP"] = well_grouped["PDSP"].fillna("")
"""

# If the barcode is a receptor, then receptor info only needs to be essiential

# List of elements to remove to isolate receptor name
elements_remove = ['-15', '-14', '-13', '-12', '-11', '-10', '-9', '-8', '-7',
                   '-6', '-5', '-4', '-3', '-2', '-1', '-0']

for index, row in well_grouped.iterrows():
    if row["Ref"] == "ref":
        unique_receptors_woNum = []
        for receptor in row["Receptor_Info"]:
            receptor_basename = receptor
            for element in elements_remove:
                receptor_basename = receptor_basename.replace(element,'')
            receptor_basename = receptor_basename.rstrip()
            unique_receptors_woNum.append(receptor_basename)

        # From this list of receptors, get acutal unique receptors
        # Get unique receptors as a sorted string
        unique_receptors = sorted(list(set(unique_receptors_woNum)))
        well_grouped.at[index, "Receptor_Info"] = unique_receptors

# Merge into the verso_worklist_df and update
verso_df = verso_df.merge(well_grouped, on="Barcode", how="left")

# Compute the length of each list in 'Receptor_Info', 1 is the minimum value
verso_df['Receptor_Length'] = verso_df['Receptor_Info'].apply(len).clip(lower=1)

# Find the maximum list length for each Row (A-H)
max_lengths = verso_df.groupby('Row')['Receptor_Length'].max().reset_index()

# Merge the max length back to the original DataFrame
verso_df = verso_df.merge(max_lengths, on='Row', suffixes=('', '_max'))

# Rename for clarity
verso_df.rename(columns={'Receptor_Length_max': 'receptor_row_max'}, inplace=True)

# Replace NaN 'Ref', PDSP, and Barcode values with an empty string
verso_df["Ref"] = verso_df["Ref"].fillna("")
verso_df["PDSP"] = verso_df["PDSP"].fillna("")
verso_df["Barcode"] = verso_df["Barcode"].replace("-1","")

# Track accumulated offset
accumulated_offset = 0
previous_row = None

# Iterate through rows
for i in range(len(verso_df)):
    current_row = verso_df.loc[i, 'Row']  # Get current row letter
    
    if previous_row is not None and current_row != previous_row:
        accumulated_offset += verso_df.loc[i - 1, 'receptor_row_max']  # Add only when row letter changes

    verso_df.loc[i, 'row_offset_sum'] = accumulated_offset  # Assign updated offset
    previous_row = current_row  # Update previous row


######################### Excel Writing Section ###############################
wb = openpyxl.load_workbook(map_template_path)
ws = wb["Sheet1"]


# Add styles and fonts
# Add arial font
arial_font_14 = openpyxl.styles.Font(name="Arial", size=14, bold=False, italic=False)
arial_font_16 = openpyxl.styles.Font(name="Arial", size=16, bold=True, italic=False)
arial_font_22 = openpyxl.styles.Font(name="Arial", size=22, bold=True, italic=False)

# Add thin border
thin_border = Side(style="thin")

# Add reference, fill color (yellow)
fill_color = openpyxl.styles.PatternFill(start_color='FDFF31',  # Hex color code for yellow
                         end_color='FDFF31',
                         fill_type="solid")

# Center alignment
center_center_align = Alignment(horizontal="center", vertical="center")

# Write Verso File Name
ws.cell(1, 5).value = verso_file_name

# Write Worklist File Name
ws.cell(2, 5).value = worklist_file_name


################## Big loop that writes all the content #######################
# Letter before loop starts is nothing
previous_letter = None
for index, row in verso_df.iterrows():
    # Establish row and column positions
    row_pos = int(row['Excel_Row']) + int(row['row_offset_sum'])
    column_pos = int(row['Excel_Column'])
    receptor_row_max = int(row['receptor_row_max'])
    cell_size = receptor_row_max + 2
    
    # Get letter of current row
    current_letter = row['Row']
    
    # Write & Merge the row letter cell iff a new row begins
    if current_letter != previous_letter:
        cell = ws.cell(row=row_pos, column=1)
        ws.merge_cells(start_row=row_pos, start_column=1, end_row=row_pos + cell_size-1, end_column=1)
        cell.value = current_letter
        cell.font = arial_font_22
        cell.alignment = center_center_align
        # Add borders
        for i in range(cell_size):
            cell = ws.cell(row=row_pos + i, column=1)
            # Write top border if first cell
            if i == 0:
                cell.border = Border(top=thin_border, left=thin_border, right=thin_border)
            
            elif i == cell_size - 1:
                cell.border = Border(bottom=thin_border, left=thin_border, right=thin_border)
            
            else:
                cell.border = Border(left=thin_border, right=thin_border)
        
    previous_letter = current_letter
    
    # Write Barcode (Convert to String to Avoid Errors)
    cell = ws.cell(row=row_pos, column=column_pos)
    cell.value = str(row['Barcode'])
    cell.font = arial_font_14

    
    # Write PDSP number (row + 1 to be one cell beneath)
    cell = ws.cell(row=row_pos + 1, column=column_pos)
    cell.value = str(row['PDSP'])
    cell.font = arial_font_16
    
    # Loop through receptors, shift rows dynamically
    for i, receptor in enumerate(row["Receptor_Info"]):
        if pd.notna(receptor):  # Avoid writing None values
            cell = ws.cell(row=row_pos + 2 + i, column=column_pos)
            cell.value = str(receptor)
            cell.font = arial_font_14
            
    # Create outer border of barcode size
    for i in range(cell_size):
        cell = ws.cell(row=row_pos + i, column=column_pos)
        # Write top border if first cell
        if i == 0:
            cell.border = Border(top=thin_border, left=thin_border, right=thin_border)
        
        elif i == cell_size - 1:
            cell.border = Border(bottom=thin_border, left=thin_border, right=thin_border)
        
        else:
            cell.border = Border(left=thin_border, right=thin_border)
    
    # If the compound is a reference, fill in yellow
    if row['Ref'] == 'ref':
        for i in range(cell_size):
            cell = ws.cell(row=row_pos + i, column=column_pos)
            cell.fill = fill_color
############################## End of Loop ####################################
    
updated_map_template_path = ""
wb.save(updated_map_template_path)
