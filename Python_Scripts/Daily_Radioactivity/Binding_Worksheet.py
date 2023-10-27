# -*- coding: utf-8 -*-
"""
Created on Tue Sep  5 14:43:49 2023

@author: TaliaAlbert
"""

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
import subprocess
import shutil
import datetime
import os
import csv
import sys
import time

print('Modules Loaded')
""" Create text files and fill them in, only progress script later"""




# Get current directory
currentdir = os.getcwd() + '\\'
inputdir = currentdir + 'input\\'
outputdir = currentdir + 'output\\'
data_filesdir = currentdir + 'data_files\\'
archivedir = currentdir + 'archive\\'

# Format the date as 'YYYY-MM-DD'
formatted_date = datetime.date.today().strftime('%Y%m%d')

log_file_path = archivedir + formatted_date + '_log.txt'
def create_directory(directory_path):
    if not os.path.exists(directory_path):
        try:
            os.makedirs(directory_path)
            print(f"Directory '{directory_path}' created.")
        except OSError as e:
            print(f"Error creating directory '{directory_path}': {e}")

def log_write(message):
    # Get the current date and time
    current_datetime = datetime.datetime.now()

    # Format the date and time as a string
    timestamp = current_datetime.strftime('%Y-%m-%d %H:%M:%S')

    # Create the log message with timestamp
    log_message = f'{timestamp} - {message}'

    # Print the log message to the console
    print(log_message)

    # Append the log message to the "log.txt" file
    with open(log_file_path, 'a') as log_file:
        log_file.write(log_message + '\n')

create_directory(inputdir)
create_directory(archivedir)
create_directory(outputdir)
log_write('Modules Loaded')

# Create and open the Worklist.txt file, unless it already exists
worklist_filename = inputdir + f'{formatted_date}_Worklist.txt'
if os.path.exists(worklist_filename):
    log_write('Worklist text file already exists')
else:
    with open(worklist_filename, 'w') as worklist_file:
        worklist_file.write('')
    
    try:
        # Open the text file with Notepad
        subprocess.Popen(['notepad.exe', worklist_filename])
    except Exception as e:
        print(f"An error occurred: {str(e)}")

# Create and open the Barcodes.txt file, unless it already exists
barcodes_filename = inputdir + f'{formatted_date}_Barcodes.txt'
if os.path.exists(barcodes_filename):
    log_write('Barcode text file already exists')
else:
    with open(barcodes_filename, 'w') as barcodes_file:
        barcodes_file.write('')
    try:
        # Open the text file with Notepad
        subprocess.Popen(['notepad.exe', barcodes_filename])
    except Exception as e:
        print(f"An error occurred: {str(e)}")




    

# Proceed input 
while True:
    user_input = input("Enter Y when ready to proceed: ").strip().lower()
    
    if user_input == 'y':
        print("Great! Proceeding...")
        break
    else:
        print("Invalid input. Please enter 'Y' to proceed.")



# Read worklist text file
with open(worklist_filename, 'r') as text_file:
    # Initialize an empty list to store the table data
    worklist_receptors = []
    
    # Read each line from the file
    for line in text_file:
        # Split the line into columns based on tabs
        columns = line.strip().split('\t')
        
        # Append the columns to the table_data list
        worklist_receptors.append(columns)
        log_write(columns)
log_write('Worklist Loaded')

# Read barcodes text file
with open(barcodes_filename, 'r') as text_file:
    # Initialize an empty list to store the table data
    barcodes = []
    
    # Read each line from the file
    for line in text_file:
        columns = line.strip()
        barcodes.append(columns)
        log_write(columns)
log_write('Barcode Loaded')

# Check that number of barcodes matches what we expect
PRIM_count = 0
SEC_count = 0
for entry in worklist_receptors:
    if entry[0].lower() == "prim":
        PRIM_count += 1
    elif entry[0].lower() == "sec":
        SEC_count += 1

total_plates = PRIM_count + 3*SEC_count

log_write('There are ' + str(total_plates) + ' plates today!')
if total_plates != len(barcodes):
    if total_plates < len(barcodes):
        log_write('More barcodes than plates were entered')
    elif total_plates > len(barcodes):
        log_write('Too few barcodes entered')
    
    sys.sleep(15)
    sys.exit(1)
   
# Read 3H-Ligand Database
Ligand_DB_path = data_filesdir + '3H-Ligand_Database.csv'
with open(Ligand_DB_path, 'r') as csvfile:
    Ligand_DB = []
    csvreader = csv.DictReader(csvfile)
    for row in csvreader:
        Ligand_DB.append(row)
log_write('3H-Ligand Database loaded: ' + Ligand_DB_path)

# Read Assay Database
Assay_DB_path = data_filesdir + 'Assay_Database.csv'
with open(Assay_DB_path, 'r') as csvfile:
    Assay_DB = []
    csvreader = csv.DictReader(csvfile)
    for row in csvreader:
        Assay_DB.append(row)
log_write('Assay Database loaded: ' + Assay_DB_path)

""" 
Create Dictionary for each Plate
Plate Name, Binding Type, Receptor, Barcode 0, Barcode 1, Barcode 2, 3H-Ligand,
Batch Number, Specific Activity, 
"""
#Create dictionary
receptors = []
for entry in worklist_receptors:
    temp_dict = {'Plate Name':entry[1].rstrip(), 'Binding Type':entry[0].rstrip()}
    receptors.append(temp_dict)

# Elucidate receptor name
for receptor in receptors:
    receptor_basename = receptor['Plate Name'].replace('-0', '').replace('-1', '').replace('-2', '').replace('-3', '').replace('-4', '').replace('-5','').replace('-6','').replace('-7','').replace('-8','').replace('-9','').replace('Rat Brain Site', '').rstrip()
    receptor.update({'Receptor':receptor_basename})
    log_write(receptor_basename)


# Match Assay Information
for receptor in receptors:
    for assay in Assay_DB:
        if receptor['Receptor'] == assay['Receptor']:
            receptor.update(assay)
log_write('Assay information matched to Receptors')

# Match 3H-Ligand Information
for receptor in receptors:
    for ligand in Ligand_DB:
        if receptor['3H-Ligand'] == ligand['3H-Ligand']:
            receptor.update(ligand)
log_write('Ligand information matched to receptors')

# Match Barcodes to plates
sec_count = 0
for index, receptor in enumerate(receptors):
    sec_index = index + (sec_count * 2)
    if receptor['Binding Type'].lower() == 'sec':
        barcode_0 = barcodes[sec_index]
        barcode_1 = barcodes[sec_index + 1]
        barcode_2 = barcodes[sec_index + 2]
        receptor.update({'Barcode 0':barcode_0, 'Barcode 1':barcode_1, 'Barcode 2':barcode_2})
        log_write(receptor['Binding Type'] + ' ' + barcode_0 + ' ' + barcode_1 + ' ' + barcode_2)
        sec_count += 1
    elif receptor['Binding Type'].lower() == 'prim':
        barcode_0 = barcodes[sec_index]
        barcode_1 = ''
        barcode_2 = ''
        receptor.update({'Barcode 0':barcode_0, 'Barcode 1':barcode_1, 'Barcode 2':barcode_2})
        log_write(receptor['Binding Type'] + ' ' + barcodes[index])
log_write('Barcodes matched to receptors')

# Change str's to float's, add Buffer Vol's, add Ligand Vol's
for receptor in receptors:
    receptor['Assay Conc. (nM)'] = float(receptor['Assay Conc. (nM)'])
    receptor['PRIM Pellet/Plate Ratio'] = float(receptor['PRIM Pellet/Plate Ratio'])
    receptor['SEC Pellet/Plate Ratio'] = float(receptor['SEC Pellet/Plate Ratio'])
    receptor['Specific Activity (Ci/mmol)'] = float(receptor['Specific Activity (Ci/mmol)'])
    if receptor['Binding Type'].lower() == 'prim':
        receptor.update({'Buffer Volume (mL)':float(5)})
        receptor.update({'Number of Plate(s)':float(1)})
        receptor.update({'Number of Pellet(s)':round(receptor['Number of Plate(s)'] * receptor['PRIM Pellet/Plate Ratio'] * 4) / 4})
    elif receptor['Binding Type'].lower() == 'sec':
        receptor.update({'Buffer Volume (mL)':float(15)})
        receptor.update({'Number of Plate(s)':float(3)})
        receptor.update({'Number of Pellet(s)':round(receptor['Number of Plate(s)'] * receptor['SEC Pellet/Plate Ratio'] * 4) / 4})

    dilution_factor = 2.5
    overage_percent = 1.44
    ligand_vol = receptor['Buffer Volume (mL)'] * receptor['Assay Conc. (nM)'] * receptor['Specific Activity (Ci/mmol)'] * (1/1000) * dilution_factor * overage_percent
    
    ############### Exceptions #######################
    # If AT2 receptor, double ligand_vol
    if receptor['Receptor'].lower() == 'at2':
        ligand_vol = ligand_vol * 2
    
    # If using GR125743, account for 1/10 Ci concentration
    if receptor['3H-Ligand'] == '3H-GR125743':
        ligand_vol = ligand_vol * 10
    ############### Exceptions #######################

    receptor.update({'Ligand Volume (uL)':ligand_vol})
    log_write(receptor['Receptor'] + ' buffer volume, ligdand volume, number of plates, and number of pellets calculated.')

"""
Gather unique receptor information and unique ligand information
"""
unique_receptors_list = []
for receptor in receptors:
    if receptor['Receptor'] not in unique_receptors_list:
        unique_receptors_list.append(receptor['Receptor'])

receptors_summary = []
for unique_receptor in unique_receptors_list:
    receptors_summary.append({'Receptor':unique_receptor})
    log_write(unique_receptor + ' is a unique receptor')

for unique_receptor in receptors_summary:
    buffer_vol = 0
    ligand_vol = 0
    plate_number = 0
    pellet_number = 0
    for receptor in receptors:
        if unique_receptor['Receptor'] == receptor['Receptor']:
            buffer_vol += receptor['Buffer Volume (mL)']
            ligand_vol += receptor['Ligand Volume (uL)']
            plate_number += receptor['Number of Plate(s)']
            pellet_number += receptor['Number of Pellet(s)']
            reference = receptor['Reference']
            ligand = receptor['3H-Ligand']
            buffer = receptor['Assay BB']
    unique_receptor.update({'Buffer Volume (mL)':buffer_vol,'Ligand Volume (uL)':round(ligand_vol,2),
                            'Number of Plates':plate_number, 'Number of Pellets':pellet_number,
                            'Reference':reference, '3H-Ligand':ligand, 'Assay BB':buffer})
    log_write(unique_receptor['Receptor'] + ' summary information determined')
log_write('Unique receptors and summary information determined')

unique_ligands_list = []
for receptor in receptors:
    if receptor['3H-Ligand'] not in unique_ligands_list:
        unique_ligands_list.append(receptor['3H-Ligand'])
unique_ligands_list = sorted(unique_ligands_list)

ligands_summary = []
for ligand in unique_ligands_list:
    ligands_summary.append({'3H-Ligand':ligand})

for ligand in ligands_summary:
    ligand_vol = 0
    for receptor in receptors:
        if ligand['3H-Ligand'] == receptor['3H-Ligand']:
            ligand_vol += receptor['Ligand Volume (uL)']
            batch_number = receptor['Batch Number']
            specific_activity = receptor['Specific Activity (Ci/mmol)']
    ligand.update({'Ligand Volume (uL)':round(ligand_vol,2), 'Batch Number':batch_number, 'Specific Activity (Ci/mmol)':specific_activity})
    log_write(ligand['3H-Ligand'] + ' summary information determined')
log_write('Unique ligands and summary information determined')

"""
Write data to Archive excel sheet
"""
archive_sheet_path = currentdir + 'Radioactivity_Archive.xlsx'
gray_switch_path = data_filesdir + 'Gray_Switch.txt'
sheet_date = datetime.date.today().strftime('%m/%d/%Y')
wb = openpyxl.load_workbook(archive_sheet_path)
ws = wb['Sheet1']

log_write('Radioactivity Archive Sheet opened: ' + archive_sheet_path)
with open(gray_switch_path, 'r') as text_file:
    for line in text_file:
        gray_switch = line.strip()

if gray_switch == '0':
    gray_switch = '1'
    cell_fill_color = 'D9D9D9'
elif gray_switch == '1':
    gray_switch = '0'
    cell_fill_color = 'FFFFFF'


columns = [
    'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
    'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
    'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD'
]

last_row = ws.max_row

border_style = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))
fill_color = PatternFill(start_color=cell_fill_color, end_color=cell_fill_color, fill_type='solid')

thick_border = Border(left=Side(style='thin'),
                      right=Side(style='thick'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

alignment = Alignment(horizontal='center')

for index, receptor in enumerate(receptors):
    row_index = index + last_row + 1
    ws.cell(row_index, 1, receptor['Binding Type'])
    ws.cell(row_index, 2, receptor['Plate Name'])
    ws.cell(row_index, 3, sheet_date)
    ws.cell(row_index, 4, receptor['Barcode 0'])
    ws.cell(row_index, 7, receptor['Barcode 1'])
    ws.cell(row_index, 10, receptor['Barcode 2'])
    ws.cell(row_index, 17, receptor['Receptor'])
    ws.cell(row_index, 18, receptor['3H-Ligand'])
    ws.cell(row_index, 19, receptor['Batch Number'])
    ws.cell(row_index, 20, receptor['Specific Activity (Ci/mmol)'])
    ws.cell(row_index, 21, receptor['Assay Conc. (nM)'])
    ws.cell(row_index, 23, receptor['Reference'])
    ws.cell(row_index, 24, receptor['Number of Plate(s)'])
    ws.cell(row_index, 25, receptor['Number of Pellet(s)'])
    ws.cell(row_index, 26, receptor['Buffer Volume (mL)'])
    ws.cell(row_index, 27, round(receptor['Ligand Volume (uL)'], 2))
    ws.cell(row_index, 28, receptor['Assay BB'])
    ws.cell(row_index, 29, receptor['PRIM Pellet/Plate Ratio'])
    ws.cell(row_index, 30, receptor['SEC Pellet/Plate Ratio'])
    
    for column in columns:
        current_cell_str = column + str(row_index) 
        current_cell = ws[current_cell_str]
        current_cell.border = border_style
        current_cell.fill = fill_color
        
        # Bold the First 2 columns
        if column == 'A' or column == 'B':
            current_cell.font = Font(bold=True)
        
        # Add Thick borders to some columns
        if column == 'C' or column == 'L' or column == 'P':
            current_cell.border = thick_border
        
        # Center Number Columns
        if column == 'E' or column == 'F' or column == 'H' or column == 'I' or column == 'K' or column == 'L' or column == 'M' or column == 'N' or column == 'O' or column == 'P'  or column == 'T' or column == 'U' or column == 'V' or column == 'X' or column == 'Y' or column == 'Z' or column == 'AA' or column == 'AC' or column == 'AD':
            current_cell.alignment = alignment

        # Add formula to automatically calculate postion in plate counter
        if column == 'E' and index == 0:
            current_cell.value = 4
        
        elif column == 'E':
            formula = '=IF(A' + str(row_index - 1) + '="SEC", K' + str(row_index - 1) + ' + 1, E' + str(row_index - 1) + ' + 1)'
            current_cell.value = formula
        
        elif column == 'H':
            formula = '=IF(A' + str(row_index) + '="SEC", E' + str(row_index) + ' + 1, "")'
            current_cell.value=formula
        
        elif column == 'K':
            formula = '=IF(A' + str(row_index) + '="SEC", H' + str(row_index) + ' + 1, "")'
            current_cell.value=formula

        
        # Add conditional formatting to make cell green when they are equal to "y" or "Y"
        if column == 'F' or column == 'I' or column == 'L' or column == 'M' or column == 'N' or column == 'O':
            green_fill = PatternFill(start_color='92d050', end_color='92d050', fill_type='solid')
            
            formula = [current_cell_str + '="y"']
            rule = FormulaRule(formula=formula, stopIfTrue=False, fill=green_fill)
            ws.conditional_formatting.add(current_cell_str, rule)
            
            formula = [current_cell_str + '="Y"']
            rule = FormulaRule(formula=formula, stopIfTrue=False, fill=green_fill)
            ws.conditional_formatting.add(current_cell_str, rule)
        
        # Make y's for secondary binding autmotacially equal eachother
        if (column == 'I' or column == 'L') and receptor['Binding Type'].lower() == 'sec':
            formula = '=F' + str(row_index)
            current_cell.value = formula
        
        # Add formula to calculate actual concentration from the counts
        if column == 'V':
            formula = '=P' + str(row_index) + '*(1/(2.22*10^12))*(1/(' + str(receptor['Specific Activity (Ci/mmol)']) + '))*(1/(0.125))*10^9'
            current_cell.value = formula
            
            number_format = '0.00'
            current_cell.number_format = number_format
        
        if index != 0:
            if receptor['3H-Ligand'] == receptors[index-1]['3H-Ligand'] and receptor['Assay Conc. (nM)'] == receptors[index-1]['Assay Conc. (nM)'] and column == 'P':
                formula = '=P' + str(row_index - 1)
                current_cell.value = formula

    log_write(receptor['Plate Name'] + ' written to Radioactivity Archive Sheet')
log_write('Radioactivity Archive Sheet written to')

wb.save(archive_sheet_path)
log_write('Radicoactivity Archive Sheet saved')

with open(gray_switch_path, 'w') as text_file:
    text_file.write(gray_switch)

"""
Move input and old output Files to archive
"""
# Move Barcode to archive
shutil.move(barcodes_filename, archivedir)
log_write('Barcode file moved to archive')

# Move Worklist to archive
shutil.move(worklist_filename, archivedir)
log_write('Worklist file moved to archive')

# Move contents of output to archive
output_files = os.listdir(outputdir)
for file in output_files:
    shutil.move(outputdir + file, archivedir)
log_write('Previous files in output directory moved to archive')



"""
Write data to binding excel sheet
"""
binding_worksheet_path = data_filesdir + 'Binding_Printout_Template.xlsx'
wb = openpyxl.load_workbook(binding_worksheet_path)
ws = wb['Sheet1']
ws.cell(2, 2, sheet_date)

log_write('Binding template loaded from: ' + binding_worksheet_path)

# Ligand Summary
for index, ligand in enumerate(ligands_summary):
    ws.cell(index + 4, 2, ligand['3H-Ligand'])
    ws.cell(index + 4, 3, ligand['Batch Number'])
    ws.cell(index + 4, 4, ligand['Specific Activity (Ci/mmol)'])
    ws.cell(index + 4, 5, ligand['Ligand Volume (uL)'])
log_write('Ligand summary populated to Binding Template')

# Pellet Summary
for index, receptor in enumerate(receptors_summary):
    ws.cell(index + 4, 7, receptor['Receptor'])
    ws.cell(index + 4, 8, receptor['Number of Pellets'])
log_write('Receptors and Pellet summary populated to Binding Template')

# Receptor information
for index, receptor in enumerate(receptors_summary):
    ws.cell(index + 14, 1, receptor['Receptor'])
    ws.cell(index + 14, 2, receptor['Assay BB'])
    ws.cell(index + 14, 3, receptor['Buffer Volume (mL)'])
    ws.cell(index + 14, 4, receptor['3H-Ligand'])
    ws.cell(index + 14, 5, receptor['Ligand Volume (uL)'])
    ws.cell(index + 14, 6, receptor['Number of Plates'])
    ws.cell(index + 14, 7, receptor['Number of Pellets'])
    ws.cell(index + 14, 8, receptor['Reference'])
log_write('Receptor information added')

# Plate Names and Barcodes
for index, receptor in enumerate(receptors):
    if index < PRIM_count and receptor['Binding Type'].lower() == 'prim':
        ws.cell(index + 4, 9, receptor['Plate Name'])
        ws.cell(index + 4, 10, index + 1)
        ws.cell(index + 4, 11, receptor['Barcode 0'])
    elif index >= PRIM_count and receptor['Binding Type'].lower() == 'sec':
        sec_index = 3*(index - PRIM_count) + PRIM_count
        ws.cell(sec_index + 4, 9, receptor['Plate Name'])
        ws.cell(sec_index + 5, 9, receptor['Plate Name'])
        ws.cell(sec_index + 6, 9, receptor['Plate Name'])
        ws.cell(sec_index + 4, 10, sec_index + 1)
        ws.cell(sec_index + 5, 10, sec_index + 2)
        ws.cell(sec_index + 6, 10, sec_index + 3)
        ws.cell(sec_index + 4, 11, receptor['Barcode 0'])
        ws.cell(sec_index + 5, 11, receptor['Barcode 1'])
        ws.cell(sec_index + 6, 11, receptor['Barcode 2'])
log_write('Barcodes and plate names added')

binding_output_path = outputdir + formatted_date + ' - Binding Printout.xlsx'
wb.save(binding_output_path)
log_write('Binding printout generated :' + binding_output_path)

log_write('Binding sheet being printed')
os.startfile(binding_output_path, 'print')
log_write('Opening Radioactivity Archive Sheet:' + archive_sheet_path)
os.startfile(archive_sheet_path)
time.sleep(15)

