# -*- coding: utf-8 -*-
import shutil
import datetime
import os
import logging
import sys
import openpyxl

parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
sys.path.insert(0, parent_dir)
import utils


# Get directories & create those which do not exist
current_dir, archive_dir, input_dir, output_dir = utils.setup_dir(create_data_files_dir='n')

# Format the date as 'YYYY-MM-DD' for tomorrow's day, unless it is Fri -> Mon
day_of_week = datetime.date.today().weekday()
if day_of_week == 4:
    formatted_date = (datetime.date.today() + datetime.timedelta(days=3)).strftime('%Y-%m-%d')
else:
    formatted_date = (datetime.date.today() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')

# Setup logging functionality
utils.setup_logging(archive_dir,log_filename=formatted_date + '_log.log')

# Get the worklist.csv file directory
message = "Please select the worklist file: "
worklist_file_dir = utils.select_file_from_input_dir(input_dir, message)

# Import the CSV into python (without header)
# Structure is Compound_Number,Receptor,Reference,Assay,PDSP_Number
worklist = utils.read_csv_file(worklist_file_dir, remove_header = True)

# Create sorted worklist by second column
worklist_sorted = sorted(worklist, key=lambda x: x[1])

# Determine if it is primary or secondary
if worklist[0][3] == 'p':
    prim_sec = 'PRIM'
    logging.info('PRIM:Primary worklist')

elif worklist[0][3] == 's':
    prim_sec = 'SEC'
    logging.info('SEC:Secondary worklist')

else:
    prim_sec = 'NA'
    logging.warning('NA:Cannot determine if PRIM or SEC')

""" Determine unique receptors """
# Get list of unique receptors (with -X afterwards)
unique_receptors_wNum = utils.unique_column_in_list(worklist,1)
logging.info('Unique receptors determined')

# List of elements to remove to isolate receptor name
elements_remove = ['-0', '-1', '-2', '-3', '-4', '-5', '-6', '-7', '-8', '-9',
                   '-10', '-11', '-12', '-13', '-14', '-15']

unique_receptors_woNum = []
for receptor in unique_receptors_wNum:
    receptor_basename = receptor
    for element in elements_remove:
        receptor_basename = receptor_basename.replace(element,'')
    receptor_basename = receptor_basename.rstrip()
    unique_receptors_woNum.append(receptor_basename)

# From this list of receptors, get acutal unique receptors
unique_receptors = sorted(list(set(unique_receptors_woNum)))

""" Create name of worklist """
# Step 1: Group receptors by prefix
grouped_receptors = {}
for receptor in unique_receptors:
    # Split at the first digit to find the prefix
    for i, char in enumerate(receptor):
        if char.isdigit():
            # Treat everything before the first digit as the prefix
            prefix = receptor[:i]
            suffix = receptor[i:]
            if prefix in grouped_receptors:
                grouped_receptors[prefix].append(suffix)
            else:
                grouped_receptors[prefix] = [suffix]
            break
    else:
        # For receptors without a digit suffix (like 'DAT')
        grouped_receptors[receptor] = []

# Step 2: Build the shortened name with spaces between groups
shortened_name = ' '.join(
    f"{prefix}{''.join(suffixes)}" if suffixes else prefix
    for prefix, suffixes in grouped_receptors.items()
)
worklist_name = formatted_date + ' ' + shortened_name + ' ' + prim_sec



""" Create csv files of unique compounds """
# move files from output into archive
utils.move_dir_files(output_dir, archive_dir)

# Determine unique compounds
unique_compounds = utils.unique_column_in_list(worklist,0)

# Determine how many plates are necessary
upper_bound = len(unique_compounds)
if upper_bound > 96:
    worklist_barcode1_dir = os.path.join(output_dir, worklist_name + ' barcode-1.csv')
    worklist_barcode2_dir = os.path.join(output_dir, worklist_name + ' barcode-2.csv')
    utils.write_list_to_csv(worklist_barcode1_dir, unique_compounds[0:96])
    utils.write_list_to_csv(worklist_barcode2_dir, unique_compounds[96:upper_bound])
    shutil.copy(worklist_file_dir, archive_dir)
    utils.move_and_rename_file(worklist_file_dir, output_dir, worklist_name + '.csv')
    
elif upper_bound <= 96:
    worklist_barcode_dir = os.path.join(output_dir, worklist_name + ' barcode.csv')
    utils.write_list_to_csv(worklist_barcode_dir, unique_compounds)
    shutil.copy(worklist_file_dir, archive_dir)
    utils.move_and_rename_file(worklist_file_dir, output_dir, worklist_name + '.csv')



""" Open Excel sheet to copy onto worklist """
google_sheet_dir = os.path.join(output_dir, worklist_name + '_google_sheet.xlsx')
wb = openpyxl.Workbook()
ws = wb.active

# Add color to PRIM/SEC column
yellow_hex_code = 'FDFF31'
green_hex_code = '6CA953'
if prim_sec == 'PRIM':
    fillable_color = green_hex_code
elif prim_sec == 'SEC':
    fillable_color = yellow_hex_code
else:
    fillable_color = '000000'

fill_color = openpyxl.styles.PatternFill(start_color=fillable_color,  # Hex color code for yellow
                         end_color=fillable_color,
                         fill_type="solid")

# Add arial font
arial_font = openpyxl.styles.Font(name="Arial", size=10, bold=False, italic=False)

# Add worklist name
ws.cell(1,1, worklist_name)
ws.cell(1,1).font = arial_font

# Populate PRIM/SEC and Plate columns
for index, plate in enumerate(unique_receptors_wNum):
    position = index + 1
    ws.cell(position, 2, prim_sec)
    ws.cell(position, 2).font = arial_font
    ws.cell(position, 2).fill = fill_color
    ws.cell(position, 3, plate)
    ws.cell(position, 3).font = arial_font

# Save excel sheet
wb.save(google_sheet_dir)

# Open excel sheet
utils.open_file(google_sheet_dir)
