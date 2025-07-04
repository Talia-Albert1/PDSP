# -*- coding: utf-8 -*-
"""
Created on Tue Sep  5 14:43:49 2023

@author: TaliaAlbert
"""
import os
import sys
import datetime
import logging
import shutil

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
import csv

import math
import time

parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
sys.path.insert(0, parent_dir)
import utils


# Get directories
current_dir, archive_dir, data_files_dir, input_dir = utils.setup_dir(create_output_dir = 'n')

# Format the date as 'YYYY-MM-DD'
formatted_date = datetime.date.today().strftime('%Y%m%d')

# Setup logging functionality
log_file_name = f'{formatted_date}_125I_bind_log.log'
utils.setup_logging(archive_dir, log_filename=log_file_name)

# Create gray_switch.txt on the first run of the script
gray_switch_dir = utils.create_file(data_files_dir, 'gray_switch.txt')
with open(gray_switch_dir, 'r+') as file:
    # Read content and remove any extra whitespace
    content = file.read().strip()
    if not content:  # If the file is blank
        file.write('1')  # Write '1' to the file
        file.flush()  # Ensure the data is written
        logging.info('Initialized file with 1')

# Get/create radioactive_waste_name txt file directory
rad_waste_name_dir = utils.create_file(data_files_dir, 'rad_waste_name.txt')
with open(rad_waste_name_dir, 'r') as file:
    rad_waste_name = file.read().strip()

# Create Archive/disposal sheets if they do not exist
archive_source_dir = os.path.join(data_files_dir, '125I-Radioactivity_Archive_blank.xlsx')
archive_destination_dir = os.path.join(current_dir, '125I-Radioactivity_Archive.xlsx')
utils.copy_and_rename(archive_source_dir, archive_destination_dir)

waste_source_dir = os.path.join(data_files_dir, '125I-Radioactive_Disposal_log_blank.xlsx')
waste_destination_dir = os.path.join(current_dir, '125I-Radioactive_Disposal_log.xlsx')
utils.copy_and_rename(waste_source_dir, waste_destination_dir)

# Tell user to close and save Radioactivity Archive xlsx sheet if it is open
logging.info('Don\'t forget to close (& save) the Radioactivity Archive Sheet before proceeding')

# Create and open the Barcodes.txt file, unless it already exists
barcodes_file_dir = os.path.join(input_dir, '125I_' + f'{formatted_date}_Barcodes.txt')
utils.create_inital_txtfile(barcodes_file_dir)

# Create and open the Worklist.txt file, unless it already exists
worklist_filename = os.path.join(input_dir, '125I_' + f'{formatted_date}_Worklist.txt')
if os.path.exists(worklist_filename):
    logging.info('Worklist text file already exists')
else:
    with open(worklist_filename, 'w') as worklist_file:
        worklist_file.write('')
    utils.open_file(worklist_filename)




    

# Proceed input 
while True:
    user_input = input("Enter Y when ready to proceed: ").strip().lower()
    
    if user_input == 'y':
        print("Great! Proceeding...")
        break
    else:
        print("Invalid input. Please enter 'Y' to proceed.")



# Read worklist text file
with open(worklist_filename, 'r') as text_file:
    # Initialize an empty list to store the table data
    worklist_receptors = []
    
    # Read each line from the file
    for line in text_file:
        # Split the line into columns based on tabs
        columns = line.strip().split('\t')
        
        # Append the columns to the table_data list
        worklist_receptors.append(columns)
        logging.info(columns)
logging.info('Worklist Loaded')

# Read barcodes text file
with open(barcodes_file_dir, 'r') as text_file:
    # Initialize an empty list to store the table data
    barcodes = []
    
    # Read each line from the file
    for line in text_file:
        columns = line.strip()
        barcodes.append(columns)
        logging.info(columns)
logging.info('Barcode Loaded')

# Remove blank lines
barcodes = list(filter(None, barcodes))
worklist_receptors = list(filter(lambda x: x != [''], worklist_receptors))

# Check that number of barcodes matches what we expect
PRIM_count = 0
SEC_count = 0
for entry in worklist_receptors:
    if entry[0].lower() == "prim":
        PRIM_count += 1
    elif entry[0].lower() == "sec":
        SEC_count += 1

total_plates = PRIM_count + 3*SEC_count

logging.info('There are ' + str(total_plates) + ' plates today!')
if total_plates != len(barcodes):
    if total_plates < len(barcodes):
        logging.info('More barcodes than plates were entered')
    elif total_plates > len(barcodes):
        logging.info('Too few barcodes entered')
    
    time.sleep(15)
    sys.exit(1)
   
# Read 125I-Ligand Database
Ligand_DB_path = os.path.join(data_files_dir, '125I-Ligand_Database.csv')
with open(Ligand_DB_path, 'r') as csvfile:
    Ligand_DB = []
    csvreader = csv.DictReader(csvfile)
    for row in csvreader:
        Ligand_DB.append(row)
logging.info('125I-Ligand Database loaded: ' + Ligand_DB_path)

# Read Assay Database
Assay_DB_path = os.path.join(data_files_dir, '125I-Assay_Database.csv')
with open(Assay_DB_path, 'r') as csvfile:
    Assay_DB = []
    csvreader = csv.DictReader(csvfile)
    for row in csvreader:
        Assay_DB.append(row)
logging.info('Assay Database loaded: ' + Assay_DB_path)

""" 
Create Dictionary for each Plate
Plate Name, Binding Type, Receptor, Barcode 0, Barcode 1, Barcode 2, 125I-Ligand,
Batch Number, Specific Activity, uCi/uL
"""
#Create dictionary
receptors = []
for entry in worklist_receptors:
    temp_dict = {'Plate Name':entry[1].rstrip(), 'Binding Type':entry[0].rstrip()}
    receptors.append(temp_dict)

# List of elements we want to replace
elements_remove = ['-0', '-1', '-2', '-3', '-4', '-5', '-6', '-7', '-8', '-9',
                   '-10', '-11', '-12', '-13', '-14', '-15', 'Rat', 'Brain',
                   'Site', 'rat', 'brain', 'site']

# Elucidate receptor name
for receptor in receptors:
    receptor_basename = receptor['Plate Name']
    for element in elements_remove:
        receptor_basename = receptor_basename.replace(element, '')
    receptor_basename = receptor_basename.rstrip()
    receptor.update({'Receptor':receptor_basename})
    logging.info(receptor_basename)


# Match Assay Information
for receptor in receptors:
    for assay in Assay_DB:
        if receptor['Receptor'] == assay['Receptor']:
            receptor.update(assay)
logging.info('Assay information matched to Receptors')

# Match 125I-Ligand Information
for receptor in receptors:
    for ligand in Ligand_DB:
        if receptor['125I-Ligand'] == ligand['125I-Ligand']:
            receptor.update(ligand)
logging.info('Ligand information matched to receptors')

# Match Barcodes to plates
sec_count = 0
for index, receptor in enumerate(receptors):
    sec_index = index + (sec_count * 2)
    if receptor['Binding Type'].lower() == 'sec':
        barcode_0 = barcodes[sec_index]
        barcode_1 = barcodes[sec_index + 1]
        barcode_2 = barcodes[sec_index + 2]
        receptor.update({'Barcode 0':barcode_0, 'Barcode 1':barcode_1, 'Barcode 2':barcode_2})
        logging.info(receptor['Binding Type'] + ' ' + barcode_0 + ' ' + barcode_1 + ' ' + barcode_2)
        sec_count += 1
    elif receptor['Binding Type'].lower() == 'prim':
        barcode_0 = barcodes[sec_index]
        barcode_1 = ''
        barcode_2 = ''
        receptor.update({'Barcode 0':barcode_0, 'Barcode 1':barcode_1, 'Barcode 2':barcode_2})
        logging.info(receptor['Binding Type'] + ' ' + barcodes[index])
logging.info('Barcodes matched to receptors')

# Change str's to float's, add Buffer Vol's
for receptor in receptors:
    receptor['Assay Conc. (nM)'] = float(receptor['Assay Conc. (nM)'])
    receptor['PRIM Pellet/Plate Ratio'] = float(receptor['PRIM Pellet/Plate Ratio'])
    receptor['SEC Pellet/Plate Ratio'] = float(receptor['SEC Pellet/Plate Ratio'])
    receptor['Specific Activity (Ci/mmol)'] = float(receptor['Specific Activity (Ci/mmol)'])
    receptor['uCi/uL'] = float(receptor['uCi/uL'])
    if receptor['Binding Type'].lower() == 'prim':
        receptor.update({'Buffer Volume (mL)':float(5)})
        receptor.update({'Number of Plate(s)':float(1)})
        receptor.update({'Number of Pellet(s)':round(receptor['Number of Plate(s)'] * receptor['PRIM Pellet/Plate Ratio'] * 8) / 8})
    elif receptor['Binding Type'].lower() == 'sec':
        receptor.update({'Buffer Volume (mL)':float(15)})
        receptor.update({'Number of Plate(s)':float(3)})
        receptor.update({'Number of Pellet(s)':round(receptor['Number of Plate(s)'] * receptor['SEC Pellet/Plate Ratio'] * 8) / 8})
    logging.info(receptor['Receptor'] + ' buffer volume, number of plates, and number of pellets calculated.')

# Calculate Decay Factor
for receptor in receptors:
    cal_date = datetime.datetime.strptime(receptor['Calibration Date'], "%m/%d/%Y").date()
    todays_date = datetime.date.today()
    days_since_cal = (todays_date - cal_date).days
    decay_factor = math.exp((-0.693/60.1)*days_since_cal)
    receptor.update({'Decay Factor':decay_factor})

# Calculate ligand vols for record keeping purposes
for receptor in receptors:
    dilution_factor = 2.5
    overage_percent = 1.44
    uCi = receptor['Buffer Volume (mL)'] * receptor['Assay Conc. (nM)'] * receptor['Specific Activity (Ci/mmol)'] * (1/1000) * dilution_factor * overage_percent 
    ligand_vol = uCi / ( receptor['uCi/uL'] * receptor['Decay Factor'])

    receptor.update({'Ligand Volume (uL)':ligand_vol})
    receptor.update({'uCi':uCi})
    logging.info(receptor['Receptor'] + ' uCi & ligand vol calculated.')

"""
Gather unique receptor information and unique ligand information
"""
unique_receptors_list = []
for receptor in receptors:
    if receptor['Receptor'] not in unique_receptors_list:
        unique_receptors_list.append(receptor['Receptor'])

receptors_summary = []
for unique_receptor in unique_receptors_list:
    receptors_summary.append({'Receptor':unique_receptor})
    logging.info(unique_receptor + ' is a unique receptor')

for unique_receptor in receptors_summary:
    buffer_vol = 0
    ligand_vol = 0
    uCi_total = 0
    plate_number = 0
    pellet_number = 0
    for receptor in receptors:
        if unique_receptor['Receptor'] == receptor['Receptor']:
            buffer_vol += receptor['Buffer Volume (mL)']
            ligand_vol += receptor['Ligand Volume (uL)']
            uCi_total += receptor['uCi']
            plate_number += receptor['Number of Plate(s)']
            pellet_number += receptor['Number of Pellet(s)']
            reference = receptor['Reference']
            ligand = receptor['125I-Ligand']
            buffer = receptor['Assay BB']
            Assay_Conc = receptor['Assay Conc. (nM)']
    unique_receptor.update({'Buffer Volume (mL)':buffer_vol,
                            'Ligand Volume (uL)':round(ligand_vol,2),
                            'uCi':round(uCi_total, 2),
                            'Number of Plates':plate_number,
                            'Number of Pellets':pellet_number,
                            'Reference':reference,
                            '125I-Ligand':ligand,
                            'Assay BB':buffer,
                            'Assay Conc. (nM)':Assay_Conc})
    logging.info(unique_receptor['Receptor'] + ' summary information determined')
logging.info('Unique receptors and summary information determined')

# Create unique ligands list
unique_ligands_list = []
for receptor in receptors:
    if receptor['125I-Ligand'] not in unique_ligands_list:
        unique_ligands_list.append(receptor['125I-Ligand'])
unique_ligands_list = sorted(unique_ligands_list)

# From the unique (and sorted) ligands list, create the ligand summary list
ligands_summary = []
for ligand in unique_ligands_list:
    ligands_summary.append({'125I-Ligand':ligand})

# Determine total ligand volume used, and batch number
for ligand in ligands_summary:
    ligand_vol = 0
    uCi_total = 0
    for receptor in receptors:
        if ligand['125I-Ligand'] == receptor['125I-Ligand']:
            ligand_vol += receptor['Ligand Volume (uL)']
            uCi_total += receptor['uCi']
            batch_number = receptor['Batch Number']
            specific_activity = receptor['Specific Activity (Ci/mmol)']
    ligand.update({'Ligand Volume (uL)':round(ligand_vol,2),
                   'uCi':round(uCi_total,2),
                   'Batch Number':batch_number,
                   'Specific Activity (Ci/mmol)':specific_activity})
    logging.info(ligand['125I-Ligand'] + ' summary information determined')
logging.info('Unique ligands and summary information determined')

# Calculate radioactive disposal information for log book
for ligand in ligands_summary:
    mCi = math.ceil(ligand['uCi'])/1000
    if mCi < 0.002:
        mCi = 0.002
    mCi_dry_waste = round(math.ceil(ligand['uCi'])*0.2)/1000
    if mCi_dry_waste < 0.001:
        mCi_dry_waste = 0.001
    mCi_sink_waste = round(mCi - mCi_dry_waste, 3)
    ligand.update({'mCi':mCi,
                   'mCi Dry Waste':mCi_dry_waste,
                   'mCi Sink Waste':mCi_sink_waste})


"""
Write data to Archive excel sheet
"""
archive_sheet_path = os.path.join(current_dir, '125I-Radioactivity_Archive.xlsx')
sheet_date = datetime.date.today().strftime('%m/%d/%Y')
wb = openpyxl.load_workbook(archive_sheet_path)
ws = wb['Sheet1']

logging.info('Radioactivity Archive Sheet opened: ' + archive_sheet_path)
with open(gray_switch_dir, 'r') as text_file:
    for line in text_file:
        gray_switch = line.strip()

if gray_switch == '0':
    gray_switch = '1'
    cell_fill_color = 'D9D9D9'
elif gray_switch == '1':
    gray_switch = '0'
    cell_fill_color = 'FFFFFF'


columns = [
    'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
    'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
    'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD',
    'AE', 'AF', 'AG', 'AH', 'AI', 'AJ'
]

last_row = ws.max_row

border_style = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))
fill_color = PatternFill(start_color=cell_fill_color, end_color=cell_fill_color, fill_type='solid')

thick_border = Border(left=Side(style='thin'),
                      right=Side(style='thick'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

alignment = Alignment(horizontal='center')

for index, receptor in enumerate(receptors):
    row_index = index + last_row + 1
    ws.cell(row_index, 1, receptor['Binding Type'])
    ws.cell(row_index, 2, receptor['Plate Name'])
    ws.cell(row_index, 3, sheet_date)
    ws.cell(row_index, 4, receptor['Barcode 0'])
    ws.cell(row_index, 7, receptor['Barcode 1'])
    ws.cell(row_index, 10, receptor['Barcode 2'])
    ws.cell(row_index, 17, receptor['Receptor'])
    ws.cell(row_index, 18, receptor['125I-Ligand'])
    ws.cell(row_index, 19, receptor['Batch Number'])
    ws.cell(row_index, 20, receptor['Specific Activity (Ci/mmol)'])
    ws.cell(row_index, 21, receptor['Assay Conc. (nM)'])
    ws.cell(row_index, 23, receptor['Reference'])
    ws.cell(row_index, 24, receptor['Number of Plate(s)'])
    ws.cell(row_index, 25, receptor['Number of Pellet(s)'])
    ws.cell(row_index, 26, receptor['Buffer Volume (mL)'])
    ws.cell(row_index, 27, round(receptor['Ligand Volume (uL)'], 2))
    ws.cell(row_index, 28, receptor['uCi/uL'])
    ws.cell(row_index, 29, round(receptor['uCi'], 2))
    ws.cell(row_index, 30, round(receptor['uCi']*0.8, 2))
    ws.cell(row_index, 31, round(receptor['uCi']*0.2, 2))
    ws.cell(row_index, 32, receptor['Assay BB'])
    ws.cell(row_index, 33, receptor['PRIM Pellet/Plate Ratio'])
    ws.cell(row_index, 34, receptor['SEC Pellet/Plate Ratio'])
    ws.cell(row_index, 35, receptor['Calibration Date'])
    ws.cell(row_index, 36, receptor['Decay Factor'])
    
    for column in columns:
        current_cell_str = column + str(row_index) 
        current_cell = ws[current_cell_str]
        current_cell.border = border_style
        current_cell.fill = fill_color
        
        # Bold the First 2 columns
        if column == 'A' or column == 'B':
            current_cell.font = Font(bold=True)
        
        # Add Thick borders to some columns
        if column == 'C' or column == 'L' or column == 'P':
            current_cell.border = thick_border
        
        # Center Number Columns
        if column == 'E' or column == 'F' or column == 'H' or column == 'I' or column == 'K' or column == 'L' or column == 'M' or column == 'N' or column == 'O' or column == 'P'  or column == 'T' or column == 'U' or column == 'V' or column == 'X' or column == 'Y' or column == 'Z' or column == 'AA' or column == 'AC' or column == 'AD':
            current_cell.alignment = alignment

        # Add formula to automatically calculate postion in plate counter
        if column == 'E' and index == 0:
            current_cell.value = 1
        
        elif column == 'E':
            formula = '=IF(A' + str(row_index - 1) + '="SEC", K' + str(row_index - 1) + ' + 1, E' + str(row_index - 1) + ' + 1)'
            current_cell.value = formula
        
        elif column == 'H':
            formula = '=IF(A' + str(row_index) + '="SEC", E' + str(row_index) + ' + 1, "")'
            current_cell.value=formula
        
        elif column == 'K':
            formula = '=IF(A' + str(row_index) + '="SEC", H' + str(row_index) + ' + 1, "")'
            current_cell.value=formula

        
        # Add conditional formatting to make cell green when they are equal to "y" or "Y"
        if column == 'F' or column == 'I' or column == 'L' or column == 'M' or column == 'N' or column == 'O':
            green_fill = PatternFill(start_color='92d050', end_color='92d050', fill_type='solid')
            
            formula = [current_cell_str + '="y"']
            rule = FormulaRule(formula=formula, stopIfTrue=False, fill=green_fill)
            ws.conditional_formatting.add(current_cell_str, rule)
            
            formula = [current_cell_str + '="Y"']
            rule = FormulaRule(formula=formula, stopIfTrue=False, fill=green_fill)
            ws.conditional_formatting.add(current_cell_str, rule)
        
        # Make y's for secondary binding autmotacially equal eachother
        if (column == 'I' or column == 'L') and receptor['Binding Type'].lower() == 'sec':
            formula = '=F' + str(row_index)
            current_cell.value = formula
        
        # Add formula to calculate actual concentration from the counts
        if column == 'V':
            formula = '=P' + str(row_index) + '*(1/(2.22*10^12))*(1/(' + str(receptor['Specific Activity (Ci/mmol)']) + '))*(1/(0.125))*10^9'
            current_cell.value = formula
            
            number_format = '0.00'
            current_cell.number_format = number_format
        
        if index != 0:
            if receptor['125I-Ligand'] == receptors[index-1]['125I-Ligand'] and receptor['Assay Conc. (nM)'] == receptors[index-1]['Assay Conc. (nM)'] and column == 'P':
                formula = '=P' + str(row_index - 1)
                current_cell.value = formula

    logging.info(receptor['Plate Name'] + ' written to Radioactivity Archive Sheet')
logging.info('Radioactivity Archive Sheet written to')

wb.save(archive_sheet_path)
logging.info('Radicoactivity Archive Sheet saved')

with open(gray_switch_dir, 'w') as text_file:
    text_file.write(gray_switch)


"""
Write data to montly sink disposal sheet
"""
radioactive_disposal_log_path = os.path.join(current_dir, '125I-Radioactive_Disposal_log.xlsx')
wb = openpyxl.load_workbook(radioactive_disposal_log_path)
ws = wb['Sheet1']
last_row = ws.max_row

for index, ligand in enumerate(ligands_summary):
    row_index = index + last_row + 1
    ws.cell(row_index, 1, sheet_date)
    ws.cell(row_index, 2, ligand['125I-Ligand'])
    ws.cell(row_index, 3, ligand['Batch Number'])
    ws.cell(row_index, 4, ligand['mCi'])
    ws.cell(row_index, 5, ligand['mCi Sink Waste'])
    ws.cell(row_index, 6, ligand['mCi Dry Waste'])
    ws.cell(row_index, 7, rad_waste_name)

wb.save(radioactive_disposal_log_path)

"""
Write data to binding excel sheet
"""
binding_worksheet_path = os.path.join(data_files_dir, '125I-Binding_Printout_Template.xlsx')
wb = openpyxl.load_workbook(binding_worksheet_path)
ws = wb['Sheet1']
ws.cell(2, 2, sheet_date)

logging.info('Binding template loaded from: ' + binding_worksheet_path)

# Ligand Summary
for index, ligand in enumerate(ligands_summary):
    ws.cell(index + 4, 2, ligand['125I-Ligand'])
    ws.cell(index + 4, 3, ligand['Batch Number'])
    ws.cell(index + 4, 4, ligand['Specific Activity (Ci/mmol)'])
    ws.cell(index + 4, 5, ligand['Ligand Volume (uL)'])
    ws.cell(index + 4, 6, ligand['mCi'])
logging.info('Ligand summary populated to Binding Template')

# Pellet Summary
for index, receptor in enumerate(receptors_summary):
    ws.cell(index + 4, 7, receptor['Receptor'])
    ws.cell(index + 4, 8, receptor['Number of Pellets'])
logging.info('Receptors and Pellet summary populated to Binding Template')

# Receptor information
start_row = 15
for index, receptor in enumerate(receptors_summary):
    ws.cell(index + start_row, 1, receptor['Receptor'])
    ws.cell(index + start_row, 2, receptor['125I-Ligand'])
    ws.cell(index + start_row, 3, receptor['Assay BB'])
    ws.cell(index + start_row, 4, receptor['Buffer Volume (mL)'])
    ws.cell(index + start_row, 5, receptor['Ligand Volume (uL)'])
    ws.cell(index + start_row, 6, receptor['Number of Plates'])
    ws.cell(index + start_row, 7, receptor['Number of Pellets'])
    ws.cell(index + start_row, 8, receptor['Reference'])
    ws.cell(index + start_row, 9, receptor['Assay Conc. (nM)'])
logging.info('Receptor information added')

# Plate Names and Barcodes
sec_index = 0
sec_count = 0
for index, receptor in enumerate(receptors):
    sec_index = index + (sec_count * 2)
    if receptor['Binding Type'].lower() == 'prim':
        ws.cell(sec_index + 4, 10, receptor['Plate Name'])
        ws.cell(sec_index + 4, 11, 'P')
        ws.cell(sec_index + 4, 12, sec_index + 1)
        ws.cell(sec_index + 4, 13, receptor['Barcode 0'])
    elif receptor['Binding Type'].lower() == 'sec':
        ws.cell(sec_index + 4, 10, receptor['Plate Name'])
        ws.cell(sec_index + 5, 10, receptor['Plate Name'])
        ws.cell(sec_index + 6, 10, receptor['Plate Name'])
        ws.cell(sec_index + 4, 11, 'S')
        ws.cell(sec_index + 5, 11, 'S')
        ws.cell(sec_index + 6, 11, 'S')
        ws.cell(sec_index + 4, 12, sec_index + 1)
        ws.cell(sec_index + 5, 12, sec_index + 2)
        ws.cell(sec_index + 6, 12, sec_index + 3)
        ws.cell(sec_index + 4, 13, receptor['Barcode 0'])
        ws.cell(sec_index + 5, 13, receptor['Barcode 1'])
        ws.cell(sec_index + 6, 13, receptor['Barcode 2'])
        sec_count += 1
logging.info('Barcodes and plate names added')

binding_output_path = os.path.join(archive_dir, formatted_date + ' - 125I-Binding Printout.xlsx')
wb.save(binding_output_path)
logging.info('Binding printout generated :' + binding_output_path)

"""
Move input files to archive
"""
# Move Barcode to archive
shutil.move(barcodes_file_dir, archive_dir)
logging.info('Barcode file moved to archive')

# Move Worklist to archive
shutil.move(worklist_filename, archive_dir)
logging.info('Worklist file moved to archive')

# Print binding sheet and open archive sheet
logging.info('Opening Radioactivity Archive Sheet:' + archive_sheet_path)
utils.open_file(archive_sheet_path)
logging.info('Binding sheet being printed')
utils.print_file(binding_output_path)
time.sleep(15)

