import logging

import pandas as pd
from pathlib import Path
import time
import datetime
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
import json

logger = logging.getLogger(__name__)

STARTING_PLATE_INDEX = 1

# column index in excel is linked to dict, where "key" is the column name --------------------------------------------------------
# in the dataframe
ARCHIVE_COLUMN_SCHEMA = {
    "A":  {"key": "Binding Type",             "bold": True,  "align": "left",   "border": "normal",      "green_toggle": False},
    "B":  {"key": "Plate Name",               "bold": True,  "align": "left",   "border": "normal",      "green_toggle": False},
    "C":  {"key": "Date",                     "bold": False, "align": "center", "border": "thick_right", "green_toggle": False},
    "D":  {"key": "Barcode 0",                "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    # plate_number
    "E":  {
        "key": None,                          "bold": False, "align": "center", "border": "normal",      "green_toggle": False,
        "formula": lambda r, idx, row, prev: 1 if idx == 0 else f'=IF(A{r - 1}="SEC", K{r - 1} + 1, E{r - 1} + 1)'
    },
    # submission_confirm
    "F":  {"key": None,                       "bold": False, "align": "center", "border": "normal",      "green_toggle": True},
    "G":  {"key": "Barcode 1",                "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    # plate_number
    "H":  {
        "key": None,                          "bold": False, "align": "center", "border": "normal",      "green_toggle": False,
        "formula": lambda r, idx, row, prev: f'=IF(A{r}="SEC", E{r} + 1, "")'
    },
    # submission_confirm
    "I":  {
        "key": None,                          "bold": False, "align": "center", "border": "normal",      "green_toggle": True,
        "formula": lambda r, idx, row, prev: f'=F{r}' if str(row.get("Binding Type")).lower() == "sec" else None
    },
    "J":  {"key": "Barcode 2",                "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    # plate_number
    "K":  {
        "key": None,                          "bold": False, "align": "center", "border": "normal",      "green_toggle": False,
        "formula": lambda r, idx, row, prev: f'=IF(A{r}="SEC", H{r} + 1, "")'
    },
    # submission_confirm
    "L":  {
        "key": None,                          "bold": False, "align": "center", "border": "thick_right", "green_toggle": True,
        "formula": lambda r, idx, row, prev: f'=F{r}' if str(row.get("Binding Type")).lower() == "sec" else None
    },
    # pellet_book_confirm
    "M":  {"key": None,                       "bold": False, "align": "center", "border": "normal",      "green_toggle": True},
    # ligand_book_confirm
    "N":  {"key": None,                       "bold": False, "align": "center", "border": "normal",      "green_toggle": True},
    # actual_counts
    "O":  {"key": None,                       "bold": False, "align": "center", "border": "normal",      "green_toggle": True},
    # worklist_confirmation
    "P":  {
        "key": None,                          "bold": False, "align": "center", "border": "thick_right", "green_toggle": False,
        "formula": lambda r, idx, row, prev: (
            f'=P{r - 1}' if (idx != 0 and prev and 
                             row.get("Ligand") == prev.get("Ligand") and 
                             row.get("Assay Conc") == prev.get("Assay Conc"))
            else None
        )
    },
    "Q":  {"key": "Receptor",                 "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    "R":  {"key": "Assay Conc",               "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    # actual_assay_conc
    "S":  {
        "key": None, "bold": False, "align": "center", "border": "normal", "green_toggle": False, "num_format": "0.00",
        "formula": lambda r, idx, row, prev: f'=P{r}*(1/2.22E+12)*(1/X{r})*(1/0.125)*10^9'
    },
    "T":  {"key": "Pellet Used",              "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    "U":  {"key": "Ligand",                   "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    "V":  {"key": "Radionuclide",             "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "W":  {"key": "Inventory Control Number", "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "X":  {"key": "Specific Activity",        "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "Y":  {"key": "Initial Quantity mCi",     "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "Z":  {"key": "Initial Volume uL",        "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AA": {"key": "uCi/uL Ratio",             "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AB": {"key": "mCi Remaining",            "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AC": {"key": "uL Remaining",             "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AD": {"key": "Date Received",            "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AE": {"key": "Date Started",             "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AF": {"key": "Date Last Used",           "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AG": {"key": "Calibration Date",         "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AH": {"key": "Decay Factor",             "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AI": {"key": "Reference",                "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    "AJ": {"key": "Number of Plates",         "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AK": {"key": "Number of Pellets",        "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AL": {"key": "Buffer Volume (mL)",       "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AM": {"key": "uL for Assay",             "bold": False, "align": "center", "border": "normal",      "green_toggle": False, "round": 2},
    "AN": {"key": "uCi for Assay",            "bold": False, "align": "center", "border": "normal",      "green_toggle": False, "round": 2},
    "AO": {"key": "Sink Waste (uCi)",         "bold": False, "align": "center", "border": "normal",      "green_toggle": False, "round": 2},
    "AP": {"key": "Dry Waste (uCi)",          "bold": False, "align": "center", "border": "normal",      "green_toggle": False, "round": 2},
    "AQ": {"key": "Assay BB",                 "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    "AR": {"key": "Filter Type",              "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AS": {"key": "Unifilter Pellet Ratio",   "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AT": {"key": "Filtermat Pellet Ratio",   "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AU": {"key": "Pellets in Inventory",     "bold": False, "align": "center", "border": "normal",      "green_toggle": False},
    "AV": {"key": "Assay DB Notes",           "bold": False, "align": "left",   "border": "normal",      "green_toggle": False},
    "AW": {"key": "Pellet DB Notes",          "bold": False, "align": "left",   "border": "normal",      "green_toggle": False}
}

PRINTOUT_COLUMN_SCHEMA = {
        "Assay Summary":{
            "df name": "Assay Summary",
            "starting row": 15,
            "borders":True,
            "schema":{
                "A": {"key": "Receptor"},
                "B": {"key": "Ligand"},
                "C": {"key": "Pellet Used"},
                "D": {"key": "Buffer"},
                "E": {"key": "Buffer Vol (mL)"},
                "F": {"key": "Ligand Vol (uL)",     "round": 2},
                "G": {"key": "# of Plates"},
                "H": {"key": "# of Pellets"},
                "I": {"key": "Filter Type"},
                "J": {"key": None},# Filter type merged with J
                "K": {"key": "Reference"},
                "L": {"key": None},#Reference Cell is merged with K
                "M": {"key": "Assay Conc. (nM)",     "round": 1}
            }
    },
        "Hot Ligand Summary":{
            "df name": "Hot Ligand Summary",
            "starting row": 4,
            "borders":False,
            "schema":{
                "B": {"key": "Ligand"},
                "C": {"key": "Inventory Control Number"},
                "D": {"key": "Specific Activity",              "round": 1},
                "E": {"key": "Ligand Vol (uL)",                "round": 2},
                "F": {"key": "Ligand in Vial (mCi)",           "format": ".3f"},
                "G": {"key": "Ligand Used (mCi)",              "format": ".3f"},
                "H": {"key": "Ligand Remaining in Vial (mCi)", "format": ".3f"},
                "I": {"key": "Finishing Vial?"}
            }
    },
        "Pellet Usage Summary":{
            "df name": "Assay Summary",
            "starting row": 4,
            "borders":False,
            "schema":{
                "J": {"key": "Receptor"},
                "K": {"key": "Pellet Used"},
                "L": {"key": "# Pellets Inventory", "round": 2},
                "M": {"key": "# of Pellets",        "round": 2},
            }
    },
        "Assay List":{
            "df name": "Assay List",
            "starting row": 4,
            "borders":True,
            "schema":{
                "N": {"key": "Plate Name"},
                "O": {"key": "Binding Type"},
                "P": {"key": None},# plate number
                "Q": {"key": "Barcode"}
            }
    }
}

def safe_load_workbook(file_path):
    """
    Safely loads an Excel workbook, prompt-retrying if the file is locked open.
    """
    while True:
        try:
            wb = openpyxl.load_workbook(file_path)
            return wb
        except PermissionError:
            logger.warning(f"Permission Denied: The file '{file_path}' is currently open or locked.")
            print(f"\n[WARNING] Please CLOSE the Excel file: {file_path}")
            user_input = input("Once closed, press [Enter] (or type 'y' and press Enter) to retry: ")
            # Optional check if you explicitly want 'y', though any key/Enter is usually more user-friendly
            time.sleep(0.5)  # Tiny buffer to let OS release file lock
        except Exception as e:
            logger.error(f"Failed to load workbook due to an unexpected error: {e}")
            raise e



def safe_save_workbook(wb, file_path):
    """
    Safely saves an Excel workbook, prompt-retrying if the file is locked open.
    """
    while True:
        try:
            wb.save(file_path)
            logger.info(f"Workbook successfully saved to {file_path}")
            return True
        except PermissionError:
            logger.warning(f"Permission Denied: Cannot save. The file '{file_path}' is currently open.")
            print(f"\n[WARNING] CRITICAL: Cannot save changes! The file is open: {file_path}")
            user_input = input("Please CLOSE the file in Excel, then press [Enter] to retry saving: ")
            time.sleep(0.5)
        except Exception as e:
            logger.error(f"Failed to save workbook due to an unexpected error: {e}")
            raise e



def write_archive_excel(
        archive_sheet_path  : Path,
        df                  : pd.DataFrame,
        column_schema       : dict[str, dict],
        user_config         : dict[str, str],
        gray_switch_name    : str,
        user_config_path    : Path,
        starting_plate_index: int = 1
):
    # ==============================================================================
    # LOAD WORKBOOK
    # ==============================================================================
    logger.info(f"Loading Radioactive Archive Sheet: {archive_sheet_path}")
    wb = safe_load_workbook(archive_sheet_path)
    ws = wb['Sheet1'] # NOTE: change this to write to different sheet ("2026", etc.) (Sheet name must be changed on excel file)

    # ==============================================================================
    # READ USER CONFIG GRAY SWITCH BOOL
    # ==============================================================================
    # Read and swap the boolean state directly inside the user_config dict ---------
    gray_switch = user_config.get(gray_switch_name, False) # default to False if missing
    if gray_switch:
        gray_switch = False
        cell_fill_color = 'D9D9D9'  # Gray
    else:
        gray_switch = True
        cell_fill_color = 'FFFFFF'  # White

    # Persist the updated boolean state back to the dictionary ---------------------
    user_config[gray_switch_name] = gray_switch

    # ==============================================================================
    # CELL STYLES
    # ==============================================================================
    border_normal = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    border_thick_right = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # this will be gray or white depending on the last script execution ------------
    fill_color = PatternFill(
        start_color=cell_fill_color,
        end_color=cell_fill_color,
        fill_type='solid'
    )
    
    # used for the y or Y confirm cells --------------------------------------------
    green_fill = PatternFill(
        start_color='92d050',
        end_color='92d050',
        fill_type='solid'
    )

    align_center = Alignment(horizontal='center')
    align_left = Alignment(horizontal='left')

    prev_row_dict = None
    last_row = ws.max_row
    # ==============================================================================
    # LOOP THROUGH DATAFRAME & COLUMN CONFIG
    # ==============================================================================
    logger.info("Writing to Archive Sheet")
    # Iterate through DataFrame using an optimized row loop
    # iterrows returns a tuple of (column index, row index), we don't need column
    for index, (df_idx, row) in enumerate(df.iterrows()): #loop through rows of the dataframe
        row_index = last_row + index + 1 #add to last rows of excel file
        row_dict = row.to_dict() #convert row from DataFrame to dict object
        
        for col_letter, config in column_schema.items(): #loop through the columns
            cell_coord = f"{col_letter}{row_index}" #get the cell coordinates "A31", "B31", etc.
            current_cell = ws[cell_coord] #get the cell
            
            # determine what values to put in the cell -----------------------------
            data_key = config["key"] # data_key is the name of the column in the dataframe
            if data_key is not None:
                val = row_dict.get(data_key)
                # Check for NaN (Pandas empty values) and convert to None for Excel
                if pd.isna(val):
                    val = None
                
                # --- DATETIME STRIPPING INTERCEPTION ---
                # Check if the value is a pandas/python datetime object or timestamp string
                elif ("date" in data_key.lower() or isinstance(val, (pd.Timestamp, datetime.datetime, datetime.date))):
                    try:
                        if isinstance(val, str):
                            # Convert string timestamp to clean date if needed
                            val = pd.to_datetime(val).strftime('%m/%d/%Y')
                        else:
                            val = val.strftime('%m/%d/%Y')
                    except Exception:
                        pass # Fallback to original value if parsing fails safely
                
                if "round" in config and val is not None:
                    val = round(val, config["round"])
                current_cell.value = val

            # Formula Extraction and Evaluation ------------------------------------
            elif "formula" in config:
                calculated_formula = config["formula"](row_index, index, row_dict, prev_row_dict)
                if calculated_formula is not None:
                    current_cell.value = calculated_formula

            # --- STARTING INDEX INTERCEPTION OVERRIDE -----------------------------
            # If it's the very first row being written, force Column E to the starting_index
            if col_letter == "E" and index == 0:
                current_cell.value = starting_plate_index

            # Apply Explicit Excel String Number Formats ---------------------------
            if "num_format" in config:
                current_cell.number_format = config["num_format"]

            # Conditional Formatting Interception ----------------------------------
            if config["green_toggle"]:
                for match_case in ['"y"', '"Y"']:
                    rule = FormulaRule(formula=[f"{cell_coord}={match_case}"], stopIfTrue=False, fill=green_fill)
                    ws.conditional_formatting.add(cell_coord, rule)

            # Cell Level Formatting Enforcement ------------------------------------
            current_cell.fill = fill_color
            current_cell.font = Font(bold=config["bold"])
            current_cell.alignment = align_center if config["align"] == "center" else align_left
            current_cell.border = border_thick_right if config["border"] == "thick_right" else border_normal

        # Cache row reference context for successive loop evaluation (Prev Row references)
        prev_row_dict = row_dict
        logger.info(f"{row_dict.get('Binding Type', 'Binding Type')} {row_dict.get('Plate Name', 'Plate')} written to Radioactivity Archive Sheet")

    logger.info('Radioactivity Archive Sheet written to successfully.')

    safe_save_workbook(wb, archive_sheet_path)
    logger.info('Radioactivity Archive Sheet saved.')


    # ==============================================================================
    # UPDATE JSON FILE
    # ==============================================================================
    # Overwrite and dump the changes straight back to your master JSON file
    try:
        with open(user_config_path, 'w') as json_file:
            json.dump(user_config, json_file, indent=4)
        logger.info(f"User configuration JSON updated successfully: {gray_switch_name} is now {gray_switch}")
    except Exception as e:
        logger.error(f"Failed to save user config JSON updates: {e}")

    logger.info("Finished Writing to Excel Archive")



def write_printout(
        printout_template_path   : Path,
        printout_dest_path       : Path,
        printout_column_schema   : dict[str, dict],
        summary_df               : dict[str, pd.DataFrame],
        now                      : datetime.datetime,
        starting_plate_index     : int=1
):
    # ==============================================================================
    # LOAD PRINTOUT TEMPLATE
    # ==============================================================================
    logger.info(f"Loading Printout Sheet: {printout_template_path}")
    wb = safe_load_workbook(printout_template_path)
    ws = wb['Sheet1']

    # ==============================================================================
    # WRITE DATE
    # ==============================================================================
    logger.info("Writing Date to Binding Printout")
    ws.cell(2, 2, now.strftime("%m/%d/%Y"))

    # ==============================================================================
    # WRITE DATAFRAMES
    # ==============================================================================
    for printout_section, dictionary in printout_column_schema.items():
        logger.info(f"Writing {printout_section} to Binding Sheet")
        df_name = dictionary.get("df name")
        starting_row = dictionary.get("starting row")
        border_bool = dictionary.get("borders")
        column_schema = dictionary.get("schema")
        df = summary_df[df_name]

        for index, (df_idx, row) in enumerate(df.iterrows()):
            row_index = starting_row + index
            row_dict = row.to_dict()
            for col_letter, config in column_schema.items():
                cell_coord = f"{col_letter}{row_index}"
                current_cell = ws[cell_coord]

                # data_key is column name in dataframe
                data_key = config["key"]
                if data_key is not None:
                    val = row_dict.get(data_key)
                    
                    if pd.isna(val):
                        val = None
                    
                    if "round" in config and val is not None:
                        val = round(val, config["round"])
                    
                    if "format" in config and val is not None:
                        val = format(val, config["format"])
                        current_cell.number_format = '@' # tells excel to treat as string
                    current_cell.value = val
                
                # add borders to some sections -------------------------------------
                if border_bool:
                    border_normal = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    current_cell.border = border_normal
                # ==================================================================
                # ASSAY SUMMARY
                # ==================================================================
                if printout_section == "Assay Summary":
                    # NOTE: The actual excel file (.xlsx) is formatted how we want
                    # NOTE: so this writing script does very little formatting
                    # NOTE: except here, where we want to alternate gray/white lines
                    if index % 2 == 0:
                        cell_fill_color = 'FFFFFF' # white for odd lines
                    else:
                        cell_fill_color = 'D9D9D9' # gray for even lines
                    fill_color = PatternFill(start_color=cell_fill_color, end_color=cell_fill_color, fill_type='solid')
                    current_cell.fill = fill_color

                # ==================================================================
                # HOTLIGAND SUMMARY
                # ==================================================================
                # we only want finishing vial to show up if we are finishing
                elif printout_section == "Hot Ligand Summary":
                    if data_key == "Finishing Vial?":
                        if val:
                            current_cell.value = "Yes"
                        else:
                            current_cell.value = None

                # ==================================================================
                # ASSAY LIST
                # ==================================================================
                # starting plate index is inserted here
                elif printout_section == "Assay List":
                    if col_letter == "P":
                        plate_number = starting_plate_index + index
                        current_cell.value = plate_number

    # ==============================================================================
    # SAVE WORKBOOK
    # ==============================================================================
    logger.info(f"Saving Workbook to: {printout_dest_path}")
    safe_save_workbook(wb, printout_dest_path)
