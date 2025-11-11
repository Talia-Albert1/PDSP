# -*- coding: utf-8 -*-
"""
Created on Tue Sep  5 14:43:49 2023

@author: TaliaAlbert
"""
# =============================================================================
# ############################## import modules ###############################
# =============================================================================
import os
import sys
import datetime
import logging
import shutil

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
import csv

import gspread
from google.oauth2.service_account import Credentials

import math
import time

parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
sys.path.insert(0, parent_dir)
import utils


# =============================================================================
# ##################### initialize directories and files ######################
# =============================================================================
# Get directories
current_dir, archive_dir, data_files_dir, input_dir = utils.setup_dir(create_output_dir = 'n')



# Format the date as 'YYYY-MM-DD'
formatted_date = datetime.date.today().strftime('%Y%m%d')



# Setup logging functionality
log_file_name = f'{formatted_date}_radioligand_log.log'
utils.setup_logging(archive_dir, log_filename=log_file_name)
logging.debug('Directories created/identified, logging setup')



# Create gray_switch.txt on the first run of the script
gray_switch_dir = utils.create_file(data_files_dir, 'gray_switch.txt')
with open(gray_switch_dir, 'r+') as file:
    # Read content and remove any extra whitespace
    content = file.read().strip()
    if not content:  # If the file is blank
        file.write('1')  # Write '1' to the file
        file.flush()  # Ensure the data is written
        logging.debug('Initialized file with 1')



# Get/create user name & initials txt file directory
user_name_txt_path = utils.create_file(data_files_dir, 'user_name.txt')
with open(user_name_txt_path, 'r') as file:
    user_name = file.read().strip()

user_initials_txt_path = utils.create_file(data_files_dir, 'user_initials.txt')
with open(user_initials_txt_path, 'r') as file:
    user_initials = file.read().strip()



# Create Archive sheet if it does not exist
archive_source_dir = os.path.join(data_files_dir, 'Radioactivity_Archive_blank.xlsx')
archive_destination_dir = os.path.join(current_dir, 'Radioactivity_Archive.xlsx')
utils.copy_and_rename(archive_source_dir, archive_destination_dir)



# Tell user to close and save Radioactivity Archive xlsx sheet if it is open
logging.info('Don\'t forget to close (& save) the Radioactivity Archive Sheet before proceeding')



# Create and open the Barcodes.txt file, unless it already exists
barcodes_txt_path = os.path.join(input_dir, f'{formatted_date}_Barcodes.txt')
utils.create_inital_txtfile(barcodes_txt_path)



# Create and open the Worklist.txt file, unless it already exists
worklist_txt_path = os.path.join(input_dir, f'{formatted_date}_Worklist.txt')
utils.create_inital_txtfile(worklist_txt_path)




    
# =============================================================================
# ################### Enter Worklist and Barcode info #########################
# =============================================================================
# Proceed input 
while True:
    user_input = input("Enter Y when ready to proceed: ").strip().lower()
    
    if user_input == 'y':
        print("Great! Proceeding...")
        break
    else:
        print("Invalid input. Please enter 'Y' to proceed.")



# Read worklist text file
with open(worklist_txt_path, 'r') as text_file:
    # Initialize an empty list to store the table data
    worklist_receptors = []
    
    # Read each line from the file
    for line in text_file:
        # Split the line into columns based on tabs
        columns = line.strip().split('\t')
        
        # Append the columns to the table_data list
        worklist_receptors.append(columns)
        logging.info(columns)
logging.info('Worklist Loaded')

# Read barcodes text file
with open(barcodes_txt_path, 'r') as text_file:
    # Initialize an empty list to store the table data
    barcodes = []
    
    # Read each line from the file
    for line in text_file:
        columns = line.strip()
        barcodes.append(columns)
        logging.info(columns)
logging.info('Barcode Loaded')

# Remove blank lines
barcodes = list(filter(None, barcodes))
worklist_receptors = list(filter(lambda x: x != [''], worklist_receptors))

# Check that number of barcodes matches what we expect
PRIM_count = 0
SEC_count = 0
for entry in worklist_receptors:
    if entry[0].lower() == "prim":
        PRIM_count += 1
    elif entry[0].lower() == "sec":
        SEC_count += 1

total_plates = PRIM_count + 3*SEC_count

logging.info('There are ' + str(total_plates) + ' plates today!')
if total_plates != len(barcodes):
    if total_plates < len(barcodes):
        logging.info('More barcodes than plates were entered')
    elif total_plates > len(barcodes):
        logging.info('Too few barcodes entered')
    
    time.sleep(15)
    sys.exit(1)


# =============================================================================
# ########################## Google Sheets ####################################
# =============================================================================
# Identify json file for google sheets
data_files_dir_files = os.listdir(data_files_dir)
for file in data_files_dir_files:
    if file.endswith('.json'):
        json_credential_name = file
json_credential_path = os.path.join(data_files_dir, json_credential_name)



# Define the scope
scope = ["https://www.googleapis.com/auth/spreadsheets",
         "https://www.googleapis.com/auth/drive"]



# Authenticate
creds = Credentials.from_service_account_file(json_credential_path, scopes=scope)
client = gspread.authorize(creds)
logging.info('google credientals authenticated')



# Google Sheet Names
google_sheet_name = 'PDSP'
google_sheet_assay_db_name = 'Assay_Param'
google_sheet_ligand_db_name = 'Hotligand_Inventory'
google_sheet_pellet_inventory_name = 'Pellet_Inventory'

# Open Sheet
sheet = client.open(google_sheet_name)

# Load assay database
worksheet = sheet.worksheet(google_sheet_assay_db_name)
assay_db = worksheet.get_all_records()
logging.debug('assay database accessed')

# load hotligand database
worksheet = sheet.worksheet(google_sheet_ligand_db_name)
ligand_db = worksheet.get_all_records()
logging.debug('hotligand database accessed')

# load pellet database
worksheet = sheet.worksheet(google_sheet_pellet_inventory_name)
pellet_inventory = worksheet.get_all_records()
logging.debug('pellet inventory accessed')


""" 
Create Dictionary for each Plate
Plate Name, Binding Type, Receptor, Barcode 0, Barcode 1, Barcode 2, Ligand,
Inventory Control Number, Specific Activity, uCi/uL
"""
#Create dictionary
# remove all spaces
receptors = []
for entry in worklist_receptors:
    temp_dict = {'Plate Name':entry[1].replace(' ', '').rstrip(), 'Binding Type':entry[0].rstrip()}
    receptors.append(temp_dict)

# List of elements we want to replace
elements_remove = ['-0', '-1', '-2', '-3', '-4', '-5', '-6', '-7', '-8', '-9',
                   '-10', '-11', '-12', '-13', '-14', '-15', 'Rat', 'Brain',
                   'Site', 'rat', 'brain', 'site', 'oxytocin', 'Oxytocin',
                   '(', ')']

# Elucidate receptor name
for receptor in receptors:
    receptor_basename = receptor['Plate Name']
    for element in elements_remove:
        receptor_basename = receptor_basename.replace(element, '')
    receptor_basename = receptor_basename.replace(' ', '').rstrip()
    receptor.update({'Receptor':receptor_basename})
    logging.debug(receptor_basename)
logging.info('Receptors Identified')

# Match Assay Information
for receptor in receptors:
    for assay in assay_db:
        assay_db_receptor_name = assay['Receptor'].replace(' ', '').rstrip()
        if receptor['Receptor'] == assay_db_receptor_name:
            receptor.update(assay)
logging.info('Assay information matched to receptors')

# Match 3H-Ligand Information, only match where current vial = TRUE
for receptor in receptors:
    for ligand in ligand_db:
        ligand_db_ligand_name = ligand['Ligand'].replace(' ','').rstrip()
        if receptor['Ligand'] == ligand_db_ligand_name and ligand['Current Vial?'] == 'TRUE':
            receptor.update(ligand)  
logging.info('Ligand information matched to receptors')

# Match pellet inventory to receptor
# some names in the inventory do not match the names from asasy DB
pellet_inventory_name_replacements = {
    'Rat P2 (BZP)':'BZP',
    'Rat P2 (PBR)':'PBR'
    }

# change names from inventory to fit the assay db
for pellet in pellet_inventory:
    if pellet['Receptor'] in pellet_inventory_name_replacements:
        pellet['Receptor'] = pellet_inventory_name_replacements[pellet['Receptor']]

# create lookup dictionary for pellet inventory
pellet_inventory_lookup = {pellet['Receptor']: pellet for pellet in pellet_inventory}

# associate pellet inventory with receptor dictionary
for receptor in receptors:
    pellet = pellet_inventory_lookup.get(receptor['Receptor'])
    if pellet:
        receptor.update({'Pellets in Inventory':pellet['Number of Pellets']})
logging.info('Pellet Inventory info matched to receptors')


# Match Barcodes to plates
sec_count = 0
for index, receptor in enumerate(receptors):
    sec_index = index + (sec_count * 2)
    if receptor['Binding Type'].lower() == 'sec':
        barcode_0 = barcodes[sec_index]
        barcode_1 = barcodes[sec_index + 1]
        barcode_2 = barcodes[sec_index + 2]
        receptor.update({'Barcode 0':barcode_0, 'Barcode 1':barcode_1, 'Barcode 2':barcode_2})
        logging.info(receptor['Binding Type'] + ' ' + barcode_0 + ' ' + barcode_1 + ' ' + barcode_2)
        sec_count += 1
    elif receptor['Binding Type'].lower() == 'prim':
        barcode_0 = barcodes[sec_index]
        barcode_1 = ''
        barcode_2 = ''
        receptor.update({'Barcode 0':barcode_0, 'Barcode 1':barcode_1, 'Barcode 2':barcode_2})
        logging.info(receptor['Binding Type'] + ' ' + barcodes[index])
logging.info('Barcodes matched to receptors')

# Change str's to float's, add Buffer Vol's
for receptor in receptors:
    receptor['Assay Conc. (nM)'] = float(receptor['Assay Conc. (nM)'])
    receptor['PRIM Pellet/Plate Ratio'] = float(receptor['PRIM Pellet/Plate Ratio'])
    receptor['SEC Pellet/Plate Ratio'] = float(receptor['SEC Pellet/Plate Ratio'])
    receptor['Specific Activity (Ci/mmol)'] = float(receptor['Specific Activity (Ci/mmol)'])
    receptor['uCi/uL Ratio'] = float(receptor['uCi/uL Ratio'])
    if receptor['Binding Type'].lower() == 'prim':
        receptor.update({'Buffer Volume (mL)':float(5)})
        receptor.update({'Number of Plate(s)':float(1)})
        receptor.update({'Number of Pellet(s)':round(receptor['Number of Plate(s)'] * receptor['PRIM Pellet/Plate Ratio'] * 8) / 8})
    elif receptor['Binding Type'].lower() == 'sec':
        receptor.update({'Buffer Volume (mL)':float(15)})
        receptor.update({'Number of Plate(s)':float(3)})
        receptor.update({'Number of Pellet(s)':round(receptor['Number of Plate(s)'] * receptor['SEC Pellet/Plate Ratio'] * 8) / 8})
    logging.debug(receptor['Receptor'] + ' buffer volume, number of plates, and number of pellets calculated.')



# Calculate ligand vols for record keeping purposes
for receptor in receptors:
    dilution_factor = 2.5
    overage_percent = 1.44
    uCi = receptor['Buffer Volume (mL)'] * receptor['Assay Conc. (nM)'] * receptor['Specific Activity (Ci/mmol)'] * (1/1000) * dilution_factor * overage_percent 
    ligand_vol = uCi / receptor['uCi/uL Ratio']

    receptor.update({'Ligand Volume (uL)':ligand_vol})
    receptor.update({'uCi':uCi})
    logging.info(receptor['Receptor'] + ' uCi & ligand vol calculated.')

"""
Gather unique receptor information and unique ligand information
"""
unique_receptors_list = []
for receptor in receptors:
    if receptor['Receptor'] not in unique_receptors_list:
        unique_receptors_list.append(receptor['Receptor'])

receptors_summary = []
for unique_receptor in unique_receptors_list:
    receptors_summary.append({'Receptor':unique_receptor})
    logging.info(unique_receptor + ' is a unique receptor')

for unique_receptor in receptors_summary:
    buffer_vol = 0
    ligand_vol = 0
    uCi_total = 0
    plate_number = 0
    pellet_number = 0
    for receptor in receptors:
        if unique_receptor['Receptor'] == receptor['Receptor']:
            buffer_vol += receptor['Buffer Volume (mL)']
            ligand_vol += receptor['Ligand Volume (uL)']
            uCi_total += receptor['uCi']
            plate_number += receptor['Number of Plate(s)']
            pellet_number += receptor['Number of Pellet(s)']
            reference = receptor['Reference']
            ligand = receptor['Ligand']
            buffer = receptor['Assay BB']
            Assay_Conc = receptor['Assay Conc. (nM)']
            pellets_in_inventory = receptor['Pellets in Inventory']
    unique_receptor.update({'Buffer Volume (mL)':buffer_vol,
                            'Ligand Volume (uL)':round(ligand_vol,2),
                            'uCi':round(uCi_total, 2),
                            'Number of Plates':plate_number,
                            'Number of Pellets':pellet_number,
                            'Reference':reference,
                            'Ligand':ligand,
                            'Assay BB':buffer,
                            'Assay Conc. (nM)':Assay_Conc,
                            'Pellets in Inventory':pellets_in_inventory})
    logging.debug(unique_receptor['Receptor'] + ' summary information determined')
logging.info('Unique receptors and summary information determined')

# Create unique ligands list
unique_ligands_list = []
for receptor in receptors:
    if receptor['Ligand'] not in unique_ligands_list:
        unique_ligands_list.append(receptor['Ligand'])
unique_ligands_list = sorted(unique_ligands_list)

# From the unique (and sorted) ligands list, create the ligand summary list
ligands_summary = []
for ligand in unique_ligands_list:
    ligands_summary.append({'Ligand':ligand})

# Determine total ligand volume used, and batch number
for ligand in ligands_summary:
    ligand_vol = 0
    uCi_total = 0
    for receptor in receptors:
        if ligand['Ligand'] == receptor['Ligand']:
            ligand_vol += receptor['Ligand Volume (uL)']
            uCi_total += receptor['uCi']
            inventory_control_number = receptor['Inventory Control Number']
            specific_activity = receptor['Specific Activity (Ci/mmol)']
            mci_in_vial = receptor['Quantity Remaining (mCi)']
            uci_ul_ratio = receptor['uCi/uL Ratio']
            radionuclide = receptor['Radio-nuclide']
            ligand.update({'Ligand Volume (uL)':round(ligand_vol,2),
                   'uCi':round(uCi_total,2),
                   'mCi in Vial':mci_in_vial,
                   'uCi/uL Ratio':uci_ul_ratio,
                   'Radionuclide':radionuclide,
                   'Inventory Control Number':inventory_control_number,
                   'Specific Activity (Ci/mmol)':specific_activity})
    logging.debug(ligand['Ligand'] + ' summary information determined')
logging.info('Unique ligands and summary information determined')

# Calculate radioactive disposal information for log book
for ligand in ligands_summary:
    mCi = math.ceil(ligand['uCi'])/1000
    if mCi < 0.002:
        mCi = 0.002
    mCi_dry_waste = round(math.ceil(ligand['uCi'])*0.2)/1000
    if mCi_dry_waste < 0.001:
        mCi_dry_waste = 0.001
    mCi_sink_waste = round(mCi - mCi_dry_waste, 3)
    ligand.update({'mCi':mCi,
                   'mCi Dry Waste':mCi_dry_waste,
                   'mCi Sink Waste':mCi_sink_waste})

# caluclate quantity of ligand remaining
for ligand in ligands_summary:
    mci_in_vial_after_use = ligand['mCi in Vial'] - ligand['mCi']
    ligand.update({'mCi Remaining in Vial':mci_in_vial_after_use})

# =============================================================================
# #################### Write data to Archive excel sheet ######################
# =============================================================================
archive_sheet_name = 'Radioactivity_Archive.xlsx'
archive_sheet_path = os.path.join(current_dir, archive_sheet_name)
sheet_date = datetime.date.today().strftime('%m/%d/%Y')
wb = openpyxl.load_workbook(archive_sheet_path)
ws = wb['Sheet1']

logging.info('Radioactivity Archive Sheet opened: ' + archive_sheet_path)
with open(gray_switch_dir, 'r') as text_file:
    for line in text_file:
        gray_switch = line.strip()

if gray_switch == '0':
    gray_switch = '1'
    cell_fill_color = 'D9D9D9'
elif gray_switch == '1':
    gray_switch = '0'
    cell_fill_color = 'FFFFFF'


columns = [
    'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
    'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
    'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD',
    'AE', 'AF', 'AG', 'AH', 'AI', 'AJ'
]

last_row = ws.max_row

border_style = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))
fill_color = PatternFill(start_color=cell_fill_color, end_color=cell_fill_color, fill_type='solid')

thick_border = Border(left=Side(style='thin'),
                      right=Side(style='thick'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

alignment = Alignment(horizontal='center')

for index, receptor in enumerate(receptors):
    row_index = index + last_row + 1
    ws.cell(row_index, 1, receptor['Binding Type'])
    ws.cell(row_index, 2, receptor['Plate Name'])
    ws.cell(row_index, 3, sheet_date)
    ws.cell(row_index, 4, receptor['Barcode 0'])
    ws.cell(row_index, 7, receptor['Barcode 1'])
    ws.cell(row_index, 10, receptor['Barcode 2'])
    ws.cell(row_index, 17, receptor['Receptor'])
    ws.cell(row_index, 18, receptor['Ligand'])
    ws.cell(row_index, 19, receptor['Radio-nuclide'])
    ws.cell(row_index, 20, receptor['Inventory Control Number'])
    ws.cell(row_index, 21, receptor['Specific Activity (Ci/mmol)'])
    ws.cell(row_index, 22, receptor['Assay Conc. (nM)'])
    ws.cell(row_index, 24, receptor['Reference'])
    ws.cell(row_index, 25, receptor['Number of Plate(s)'])
    ws.cell(row_index, 26, receptor['Number of Pellet(s)'])
    ws.cell(row_index, 27, receptor['Buffer Volume (mL)'])
    ws.cell(row_index, 28, round(receptor['Ligand Volume (uL)'], 2))
    ws.cell(row_index, 29, receptor['uCi/uL Ratio'])
    ws.cell(row_index, 30, round(receptor['uCi'], 2))
    ws.cell(row_index, 31, round(receptor['uCi']*0.8, 2))
    ws.cell(row_index, 32, round(receptor['uCi']*0.2, 2))
    ws.cell(row_index, 33, receptor['Assay BB'])
    ws.cell(row_index, 34, receptor['PRIM Pellet/Plate Ratio'])
    ws.cell(row_index, 35, receptor['SEC Pellet/Plate Ratio'])
    ws.cell(row_index, 36, receptor['Pellets in Inventory'])
    
    for column in columns:
        current_cell_str = column + str(row_index) 
        current_cell = ws[current_cell_str]
        current_cell.border = border_style
        current_cell.fill = fill_color
        
        # Bold the First 2 columns
        if column == 'A' or column == 'B':
            current_cell.font = Font(bold=True)
        
        # Add Thick borders to right side of some columns
        if column == 'C' or column == 'L' or column == 'P':
            current_cell.border = thick_border
        
        # Center Number Columns
        if column == 'E' or column == 'F' or column == 'H' or column == 'I' or column == 'K' or column == 'L' or column == 'M' or column == 'N' or column == 'O' or column == 'P'  or column == 'T' or column == 'U' or column == 'V' or column == 'W' or column == 'X' or column == 'Y' or column == 'Z' or column == 'AA' or column == 'AB' or column == 'AC' or column == 'AD':
            current_cell.alignment = alignment

        # Add formula to automatically calculate postion in plate counter
        if column == 'E' and index == 0:
            current_cell.value = 4
        
        elif column == 'E':
            formula = '=IF(A' + str(row_index - 1) + '="SEC", K' + str(row_index - 1) + ' + 1, E' + str(row_index - 1) + ' + 1)'
            current_cell.value = formula
        
        elif column == 'H':
            formula = '=IF(A' + str(row_index) + '="SEC", E' + str(row_index) + ' + 1, "")'
            current_cell.value=formula
        
        elif column == 'K':
            formula = '=IF(A' + str(row_index) + '="SEC", H' + str(row_index) + ' + 1, "")'
            current_cell.value=formula

        
        # Add conditional formatting to make cell green when they are equal to "y" or "Y"
        if column == 'F' or column == 'I' or column == 'L' or column == 'M' or column == 'N' or column == 'O':
            green_fill = PatternFill(start_color='92d050', end_color='92d050', fill_type='solid')
            
            formula = [current_cell_str + '="y"']
            rule = FormulaRule(formula=formula, stopIfTrue=False, fill=green_fill)
            ws.conditional_formatting.add(current_cell_str, rule)
            
            formula = [current_cell_str + '="Y"']
            rule = FormulaRule(formula=formula, stopIfTrue=False, fill=green_fill)
            ws.conditional_formatting.add(current_cell_str, rule)
        
        # Make y's for secondary binding autmotacially equal eachother
        if (column == 'I' or column == 'L') and receptor['Binding Type'].lower() == 'sec':
            formula = '=F' + str(row_index)
            current_cell.value = formula
        
        # Add formula to calculate actual concentration from the counts
        if column == 'W':
            formula = '=P' + str(row_index) + '*(1/(2.22*10^12))*(1/(' + str(receptor['Specific Activity (Ci/mmol)']) + '))*(1/(0.125))*10^9'
            current_cell.value = formula
            
            number_format = '0.00'
            current_cell.number_format = number_format
        
        if index != 0:
            if receptor['Ligand'] == receptors[index-1]['Ligand'] and receptor['Assay Conc. (nM)'] == receptors[index-1]['Assay Conc. (nM)'] and column == 'P':
                formula = '=P' + str(row_index - 1)
                current_cell.value = formula

    logging.info(receptor['Plate Name'] + ' written to Radioactivity Archive Sheet')
logging.info('Radioactivity Archive Sheet written to')

wb.save(archive_sheet_path)
logging.info('Radicoactivity Archive Sheet saved')

with open(gray_switch_dir, 'w') as text_file:
    text_file.write(gray_switch)


# =============================================================================
# ################## Write Rad Waste Data to Google Sheets ####################
# =============================================================================
google_sheet_rad_waste_log_name = 'Hotligand_Log'

# create list of lists hot waste to write to google sheet, done in 1 api call
rad_waste_list = []
for ligand in ligands_summary:
    rad_waste_list.append([sheet_date, ligand['Ligand'], ligand['Radionuclide'],
                           ligand['Inventory Control Number'], ligand['mCi'],
                           ligand['mCi Sink Waste'], ligand['mCi Dry Waste'],
                           user_name])

worksheet = sheet.worksheet(google_sheet_rad_waste_log_name)
worksheet.append_rows(rad_waste_list,value_input_option="USER_ENTERED")

# =============================================================================
# ############ Write pellet usage data to google sheets #######################
# =============================================================================
google_sheet_pellet_log_name = 'Pellet_Log'

pellet_log_list = []
for receptor in receptors_summary:
    pellets_used = -1 * receptor['Number of Pellets']
    pellet_log_list.append([sheet_date, receptor['Receptor'],
                            user_initials, pellets_used])

worksheet = sheet.worksheet(google_sheet_pellet_log_name)
worksheet.append_rows(pellet_log_list,value_input_option="USER_ENTERED")


# =============================================================================
# ############## Write data to binding excel printout sheet ###################
# =============================================================================
binding_worksheet_path = os.path.join(data_files_dir, 'Binding_Printout_Template.xlsx')
wb = openpyxl.load_workbook(binding_worksheet_path)
ws = wb['Sheet1']
ws.cell(2, 2, sheet_date)

logging.info('Binding template loaded from: ' + binding_worksheet_path)

# Ligand Summary
for index, ligand in enumerate(ligands_summary):
    ws.cell(index + 4, 2, ligand['Ligand'])
    ws.cell(index + 4, 3, ligand['Inventory Control Number'])
    ws.cell(index + 4, 4, ligand['Specific Activity (Ci/mmol)'])
    ws.cell(index + 4, 5, ligand['Ligand Volume (uL)'])
    ws.cell(index + 4, 6, format(ligand['mCi in Vial'], ".3f"))
    ws.cell(index + 4, 7, format(ligand['mCi'], ".3f"))
    ws.cell(index + 4, 8, format(ligand['mCi Remaining in Vial'], ".3f"))
logging.info('Ligand summary populated to Binding Template')

# Pellet Summary
for index, receptor in enumerate(receptors_summary):
    ws.cell(index + 4, 9, receptor['Receptor'])
    ws.cell(index + 4, 10, receptor['Pellets in Inventory'])
    ws.cell(index + 4, 11, receptor['Number of Pellets'])
logging.info('Receptors and Pellet summary populated to Binding Template')

# Receptor information
start_row = 15
for index, receptor in enumerate(receptors_summary):
    if index % 2 == 0:
        # white for odd lines
        cell_fill_color = 'FFFFFF'
    else:
        # gray for even lines
        cell_fill_color = 'D9D9D9'
    fill_color = PatternFill(start_color=cell_fill_color, end_color=cell_fill_color, fill_type='solid')
    
    cell_receptor = ws.cell(index + start_row, 1)
    cell_receptor.value = receptor['Receptor']
    cell_receptor.fill = fill_color
    cell_receptor.border = border_style
    
    cell_ligand = ws.cell(index + start_row, 2)
    cell_ligand.value = receptor['Ligand']
    cell_ligand.fill = fill_color
    cell_ligand.border = border_style
    
    cell_assay_bb = ws.cell(index + start_row, 3)
    cell_assay_bb.value = receptor['Assay BB']
    cell_assay_bb.fill = fill_color
    cell_assay_bb.border = border_style
    
    cell_buff_vol = ws.cell(index + start_row, 4)
    cell_buff_vol.value = receptor['Buffer Volume (mL)']
    cell_buff_vol.fill = fill_color
    cell_buff_vol.border = border_style
    
    cell_ligand_vol = ws.cell(index + start_row, 5)
    cell_ligand_vol.value = receptor['Ligand Volume (uL)']
    cell_ligand_vol.fill = fill_color
    cell_ligand_vol.border = border_style
    
    cell_num_of_plates = ws.cell(index + start_row, 6)
    cell_num_of_plates.value = receptor['Number of Plates']
    cell_num_of_plates.fill = fill_color
    cell_num_of_plates.border = border_style
    
    cell_num_of_pellets = ws.cell(index + start_row, 7)
    cell_num_of_pellets.value = receptor['Number of Pellets']
    cell_num_of_pellets.fill = fill_color
    cell_num_of_pellets.border = border_style
    
    cell_reference = ws.cell(index + start_row, 8)
    cell_reference.value = receptor['Reference']
    cell_reference.fill = fill_color
    cell_reference.border = border_style
    ws.cell(index + start_row, 9).border = border_style
    ws.cell(index + start_row, 10).border = border_style
    

    cell_assay_conc = ws.cell(index + start_row, 11)
    cell_assay_conc.value = receptor['Assay Conc. (nM)']
    cell_assay_conc.fill = fill_color
    cell_assay_conc.border = border_style

logging.info('Receptor information added')

# Plate Names and Barcodes
sec_index = 0
sec_count = 0

# format left, middle, and right borders for rows
border_left = Border(left=Side(style='thin'))
border_right = Border(right=Side(style='thin'))
border_middle = Border()

for index, receptor in enumerate(receptors):
    sec_index = index + (sec_count * 2)
    
    # format final rows border differently, add thin line on bottom
    if index == len(receptors)-1:
        border_final_left = Border(left=Side(style='thin'),
                              bottom=Side(style='thin'))
        border_final_right = Border(right=Side(style='thin'),
                              bottom=Side(style='thin'))
        border_final_middle = Border(bottom=Side(style='thin'))
        
    else:
        border_final_left = border_left
        border_final_right = border_right
        border_final_middle = border_middle

    if receptor['Binding Type'].lower() == 'prim':
        ws.cell(sec_index + 4, 12, receptor['Plate Name']).border = border_final_left
        ws.cell(sec_index + 4, 13, 'P').border = border_final_middle
        ws.cell(sec_index + 4, 14, sec_index + 4).border = border_final_middle
        ws.cell(sec_index + 4, 15, receptor['Barcode 0']).border = border_final_right
    elif receptor['Binding Type'].lower() == 'sec':
        ws.cell(sec_index + 4, 12, receptor['Plate Name']).border = border_left
        ws.cell(sec_index + 5, 12, receptor['Plate Name']).border = border_left
        ws.cell(sec_index + 6, 12, receptor['Plate Name']).border = border_final_left
        ws.cell(sec_index + 4, 13, 'S').border = border_middle
        ws.cell(sec_index + 5, 13, 'S').border = border_middle
        ws.cell(sec_index + 6, 13, 'S').border = border_final_middle
        ws.cell(sec_index + 4, 14, sec_index + 4).border = border_middle
        ws.cell(sec_index + 5, 14, sec_index + 5).border = border_middle
        ws.cell(sec_index + 6, 14, sec_index + 6).border = border_final_middle
        ws.cell(sec_index + 4, 15, receptor['Barcode 0']).border = border_right
        ws.cell(sec_index + 5, 15, receptor['Barcode 1']).border = border_right
        ws.cell(sec_index + 6, 15, receptor['Barcode 2']).border = border_final_right
        sec_count += 1
logging.info('Barcodes and plate names added')

binding_output_path = os.path.join(archive_dir, formatted_date + ' - Binding Printout.xlsx')
wb.save(binding_output_path)
logging.info('Binding printout generated :' + binding_output_path)




# =============================================================================
# #################### Move input files to archive ############################
# =============================================================================
# Move Barcode to archive
shutil.move(barcodes_txt_path, archive_dir)
logging.info('Barcode file moved to archive')

# Move Worklist to archive
shutil.move(worklist_txt_path, archive_dir)
logging.info('Worklist file moved to archive')

# Print binding sheet and open archive sheet
logging.info('Opening Radioactivity Archive Sheet:' + archive_sheet_path)
utils.open_file(archive_sheet_path)
logging.info('Binding sheet being printed')
utils.print_file(binding_output_path)
time.sleep(15)

