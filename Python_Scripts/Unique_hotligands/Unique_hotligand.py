# -*- coding: utf-8 -*-
import os
import openpyxl

# Get current directory
currentdir = os.getcwd() + '\\'

master_list = currentdir + 'Receptor_hotligand.txt'
excel_file = currentdir + 'Receptor_hotligand.xlsx'

# Read text file
with open(master_list, 'r') as text_file:
    # Initialize an empty list to store the table data
    receptors_hotligands = []
    
    # Read each line from the file
    for line in text_file:
        # Split the line into columns based on tabs
        columns = line.strip().split('\t')
        
        # Append the columns to the table_data list
        receptors_hotligands.append(columns)
    
# Get unique hot ligands
unique_hotligands = []
for ligand in receptors_hotligands:
    if ligand[1] not in unique_hotligands:
        unique_hotligands.append(ligand[1])

# Sort this unique list alphabetically
unique_hotligands.sort()

# Total matches list is where our final result will live
total_matches = []
for ligand in unique_hotligands:
    # Empty matches at the start of every for loop
    matches = []
    for receptor in receptors_hotligands:
        if ligand == receptor[1]:
            matches.append(receptor[0])
    # Easier to work with a list where 0 is hot ligand and 1 is a comma separated string of receptors
    matches_list = [', '.join(matches)]
    ligand = [ligand]
    total_matches.append(ligand + matches_list)

# Write output to excel file
wb = openpyxl.Workbook()
wb.save(excel_file)
wb = openpyxl.load_workbook(excel_file)
ws = wb['Sheet']

for index, ligand in enumerate(total_matches):
    ws.cell(index+1, 1, ligand[0])
    ws.cell(index+1, 2, ligand[1])
wb.save(excel_file) 